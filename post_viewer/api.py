from __future__ import annotations

import json
import base64
import hashlib
import html as html_lib
import csv
import io
import os
import re
import socket
import zipfile
from pathlib import Path
from typing import Any, Callable
from datetime import date, datetime, timedelta, timezone
from xml.etree import ElementTree as ET
from urllib.parse import parse_qs, quote, urlencode, urlparse, urlunparse
from urllib.request import Request as UrlRequest, urlopen

from fastapi import FastAPI, File, HTTPException, Query, Request as FastAPIRequest, UploadFile
from fastapi.responses import HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


POST_TABLE = "t_fund_generated_posts"
PUBLISH_TASK_TABLE = "post_supplement_lib.t_publish_tasks"
DIMENSION_CATEGORY_TABLE = "t_content_dimension_categories"
DIMENSION_OPTION_TABLE = "t_content_dimension_options"
TASK_SUBMISSION_TABLE = "crawler_app.task_submissions"
SETTLEMENT_TABLE = "crawler_app.kol_business_settlements"
KOL_DAILY_METRICS_TABLE = "crawler_app.kol_daily_metrics"
KOL_BASE_PROFILE_TABLE = "crawler_app.kol_base_profiles"
ALIPAY_HOT_FUND_RANKINGS_TABLE = "crawler_app.alipay_hot_fund_rankings"
KOL_MAX_LIMIT = 2000
KOL_READ_SOURCE_DEFAULT_URL = "https://docs.qq.com/sheet/DYnBmZ2drS3B2QVRZ?tab=BB08J2"
KOL_READ_SOURCE_PLATFORM = "理财通"
KOL_READ_SOURCE_TYPE = "内部"
KOL_SORT_OPTIONS = {
    "base_id": "CASE WHEN b.id IS NULL THEN 1 ELSE 0 END, b.id ASC, m.metric_date DESC, m.platform ASC, m.kol_name ASC",
    "title": "m.kol_name ASC, m.metric_date DESC",
    "title_desc": "m.kol_name DESC, m.metric_date DESC",
    "date_desc": "m.metric_date DESC, m.platform ASC, b.group_name ASC, m.kol_name ASC",
    "date_asc": "m.metric_date ASC, m.platform ASC, b.group_name ASC, m.kol_name ASC",
    "platform": "m.platform ASC, b.group_name ASC, m.kol_name ASC, m.metric_date DESC",
    "group": "b.group_name ASC, m.kol_name ASC, m.metric_date DESC",
    "fans_desc": "m.fans_count DESC, m.metric_date DESC, m.kol_name ASC",
    "growth_desc": "m.growth_count DESC, m.metric_date DESC, m.kol_name ASC",
    "read_desc": "m.read_count DESC, m.metric_date DESC, m.kol_name ASC",
}
KOL_MISSING_OPTIONS = {
    "": "全部",
    "fans_empty": "粉丝数为空",
    "growth_empty": "增粉数为空",
    "fans_or_growth_empty": "粉丝数或增粉数为空",
    "fans_and_growth_empty": "粉丝数和增粉数都为空",
}


def _viewer_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _capture_mounts() -> list[tuple[str, Path]]:
    configured = os.environ.get("EASY_VIEWER_CAPTURE_ROOT", "").strip()
    candidates: list[tuple[str, Path]] = []
    if configured:
        configured_paths = [Path(item.strip()) for item in configured.split(os.pathsep) if item.strip()]
        candidates.extend((f"/captures-extra-{index}" if index > 1 else "/captures", path) for index, path in enumerate(configured_paths, start=1))
    root = _viewer_root()
    finance_captures = root.parent / "adb" / "apps" / "finance_crawler" / "captures"
    candidates.extend(
        [
            ("/captures", root / "captures"),
            ("/captures", finance_captures),
            ("/captures-adb-finance", finance_captures),
            ("/captures-adb-runtime", root.parent / "adb" / "runtime" / "captures"),
            ("/captures-adb-tmp", root.parent / "adb" / "tmp"),
            ("/captures-adb-exports", root.parent / "adb" / "exports"),
        ]
    )
    mounts: list[tuple[str, Path]] = []
    seen_mounts: set[str] = set()
    for mount_path, path in candidates:
        try:
            resolved = path.resolve()
            if not path.exists() or not path.is_dir():
                continue
            if mount_path in seen_mounts:
                mount_path = f"{mount_path}-{len(seen_mounts) + 1}"
            mounts.append((mount_path, path))
            seen_mounts.add(mount_path)
        except OSError:
            continue
    return mounts


def _capture_root() -> Path | None:
    mounts = _capture_mounts()
    return mounts[0][1] if mounts else None


def _capture_url_path(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    try:
        source = Path(text).resolve()
    except OSError:
        return ""
    for mount_path, capture_root in _capture_mounts():
        try:
            relative = source.relative_to(capture_root.resolve())
        except (OSError, ValueError):
            continue
        return mount_path.rstrip("/") + "/" + quote(relative.as_posix())
    return ""


def _load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if not os.environ.get(key):
            os.environ[key] = value


def _load_default_env() -> None:
    _load_env_file(_viewer_root() / ".env")
    _load_env_file(_viewer_root().parent / "easy-flow" / ".env")


def _safe_identifier(value: str) -> str:
    if not re.fullmatch(r"[A-Za-z0-9_]+", value or ""):
        raise RuntimeError(f"unsafe MySQL identifier: {value!r}")
    return value


def _mysql_config() -> dict[str, Any]:
    _load_default_env()
    required = ["MYSQL_HOST", "MYSQL_PORT", "MYSQL_USER", "MYSQL_PASSWORD", "MYSQL_DATABASE"]
    missing = [key for key in required if not os.environ.get(key)]
    if missing:
        raise RuntimeError("missing MySQL env keys: " + ", ".join(missing))
    return {
        "host": os.environ["MYSQL_HOST"],
        "port": int(os.environ.get("MYSQL_PORT", "3306")),
        "user": os.environ["MYSQL_USER"],
        "password": os.environ["MYSQL_PASSWORD"],
        "database": _safe_identifier(os.environ["MYSQL_DATABASE"]),
        "charset": "utf8mb4",
        "cursorclass": _dict_cursor_class(),
        "connect_timeout": int(os.environ.get("MYSQL_CONNECT_TIMEOUT", "30")),
        "read_timeout": int(os.environ.get("MYSQL_READ_TIMEOUT", "30")),
        "write_timeout": int(os.environ.get("MYSQL_WRITE_TIMEOUT", "30")),
    }


def _dict_cursor_class() -> Any:
    try:
        import pymysql
    except ImportError as exc:
        raise RuntimeError("PyMySQL is required. Install with: pip install -e .") from exc
    return pymysql.cursors.DictCursor


def _connect() -> Any:
    try:
        import pymysql
    except ImportError as exc:
        raise RuntimeError("PyMySQL is required. Install with: pip install -e .") from exc
    return pymysql.connect(**_mysql_config())


def _json_value(value: Any, fallback: Any) -> Any:
    if value is None:
        return fallback
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, bytes):
        value = value.decode("utf-8")
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return fallback
        try:
            return json.loads(text)
        except json.JSONDecodeError:
            return fallback
    return fallback


def _first_present(*values: Any) -> Any:
    for value in values:
        if value not in (None, ""):
            return value
    return ""


LOCAL_URL_HOSTS = {"127.0.0.1", "localhost", "0.0.0.0", "::1"}


def _netloc_hostname(netloc: str) -> str:
    return (urlparse(f"//{netloc}").hostname or "").lower()


def _netloc_port(netloc: str) -> str:
    port = urlparse(f"//{netloc}").port
    return f":{port}" if port else ""


def _configured_public_base(default_scheme: str) -> tuple[str, str] | None:
    public_url = os.environ.get("EASY_VIEWER_PUBLIC_BASE_URL", "").strip()
    if public_url:
        parsed = urlparse(public_url)
        if parsed.scheme in {"http", "https"} and parsed.netloc:
            return parsed.scheme, parsed.netloc
    public_host = os.environ.get("EASY_VIEWER_PUBLIC_HOST", "").strip()
    if public_host:
        return default_scheme, public_host
    return None


def _is_lan_ip(value: str) -> bool:
    if not value or value.startswith("127.") or value.startswith("169.254."):
        return False
    return value.startswith("10.") or value.startswith("192.168.") or re.fullmatch(r"172\.(1[6-9]|2\d|3[0-1])\..+", value) is not None


def _detect_lan_ip() -> str:
    candidates: list[str] = []
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
            sock.connect(("8.8.8.8", 80))
            candidates.append(sock.getsockname()[0])
    except OSError:
        pass
    try:
        candidates.extend(socket.gethostbyname_ex(socket.gethostname())[2])
    except OSError:
        pass
    for candidate in candidates:
        if _is_lan_ip(candidate):
            return candidate
    return ""


def _request_public_base(request: FastAPIRequest | None) -> tuple[str, str] | None:
    if request is None:
        return None
    host = request.headers.get("x-forwarded-host") or request.headers.get("host") or request.url.netloc
    if not host:
        return None
    scheme = request.headers.get("x-forwarded-proto") or request.url.scheme or "http"
    scheme = scheme.split(",", 1)[0].strip()
    host = host.split(",", 1)[0].strip()
    configured = _configured_public_base(scheme)
    if configured:
        return configured
    if _netloc_hostname(host) in LOCAL_URL_HOSTS:
        lan_ip = _detect_lan_ip()
        if lan_ip:
            return scheme, f"{lan_ip}{_netloc_port(host)}"
    return scheme, host


def _externalize_local_url(value: Any, request: FastAPIRequest | None) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    parsed = urlparse(text)
    if parsed.scheme not in {"http", "https"}:
        capture_path = _capture_url_path(text)
        if not capture_path:
            return text
        public_base = _request_public_base(request) or _configured_public_base("http")
        if public_base is None:
            return capture_path
        scheme, host = public_base
        return f"{scheme}://{host}{capture_path}"
    hostname = (parsed.hostname or "").lower()
    if hostname not in LOCAL_URL_HOSTS:
        return text
    public_base = _request_public_base(request)
    if public_base is None:
        return text
    scheme, host = public_base
    return urlunparse(parsed._replace(scheme=scheme, netloc=host))


def _deep_find_value(value: Any, keys: set[str]) -> Any:
    if isinstance(value, dict):
        for key, item in value.items():
            if str(key) in keys and item not in (None, ""):
                return item
        for item in value.values():
            found = _deep_find_value(item, keys)
            if found not in (None, ""):
                return found
    elif isinstance(value, list):
        for item in value:
            found = _deep_find_value(item, keys)
            if found not in (None, ""):
                return found
    return ""


def _date_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return text[:10] if len(text) >= 10 else text


def _style_label(style: str) -> str:
    return {
        "rant": "吐槽共鸣",
        "question": "提问互动",
        "analysis": "轻分析",
    }.get(style, style or "未知")


def _row_to_post(row: dict[str, Any]) -> dict[str, Any]:
    source = _json_value(row.get("source_json"), {})
    raw_post = _json_value(row.get("raw_post_json"), {})
    style = str(source.get("style") or (raw_post.get("source") or {}).get("style") or "unknown")
    return {
        "content_id": str(row.get("content_id") or ""),
        "rank": int(row.get("rank_no") or 0),
        "score": float(row.get("score") or 0),
        "style": style,
        "style_label": _style_label(style),
        "dimension_label": str(row.get("dimension_label") or ""),
        "title": str(row.get("title") or ""),
        "body": str(row.get("body") or ""),
        "hashtags": [str(item) for item in _json_value(row.get("hashtags_json"), [])],
        "disclaimer": str(row.get("disclaimer") or ""),
        "status": str(row.get("status") or ""),
        "trade_date": str(row.get("trade_date") or ""),
        "generated_at": str(row.get("generated_at") or ""),
        "run_id": str(row.get("run_id") or ""),
        "human_score": None,
        "quality_passed": None,
        "quality_issues": [],
        "quality_warnings": [],
    }


def _latest_batch_filter(connection: Any) -> tuple[str, tuple[Any, ...]]:
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT trade_date, generated_at, run_id
            FROM {POST_TABLE}
            ORDER BY trade_date DESC, generated_at DESC, run_id DESC
            LIMIT 1
            """
        )
        row = cursor.fetchone()
    if not row:
        return "", ()
    return "WHERE trade_date = %s AND generated_at = %s AND run_id = %s", (
        row["trade_date"],
        row["generated_at"],
        row["run_id"],
    )


def _batch_filter(connection: Any, trade_date: str = "", run_id: str = "", generated_at: str = "") -> tuple[str, tuple[Any, ...]]:
    clauses: list[str] = []
    params: list[Any] = []
    if trade_date:
        clauses.append("trade_date = %s")
        params.append(trade_date)
    if run_id:
        clauses.append("run_id = %s")
        params.append(run_id)
    if generated_at:
        clauses.append("generated_at = %s")
        params.append(generated_at)
    if clauses:
        return "WHERE " + " AND ".join(clauses), tuple(params)
    return _latest_batch_filter(connection)


def _posts_payload(*, trade_date: str = "", run_id: str = "", generated_at: str = "") -> dict[str, Any]:
    with _connect() as connection:
        where_sql, params = _batch_filter(connection, trade_date=trade_date, run_id=run_id, generated_at=generated_at)
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT trade_date, generated_at, run_id, content_id, rank_no, score,
                       dimension_label, title, body, disclaimer, hashtags_json,
                       signals_json, source_json, raw_post_json, status
                FROM {POST_TABLE}
                {where_sql}
                ORDER BY rank_no ASC, content_id ASC
                """,
                params,
            )
            rows = cursor.fetchall()
    posts = [_row_to_post(row) for row in rows]
    style_counts: dict[str, int] = {}
    for post in posts:
        style_counts[post["style"]] = style_counts.get(post["style"], 0) + 1
    first = rows[0] if rows else {}
    signals = _json_value(first.get("signals_json") if first else None, {})
    return {
        "source": "mysql",
        "table": POST_TABLE,
        "run_id": str(first.get("run_id") or ""),
        "trade_date": str(first.get("trade_date") or ""),
        "generated_at": str(first.get("generated_at") or ""),
        "generated_count": len(posts),
        "quality_passed": True,
        "quality_issues": [],
        "quality_warnings": [],
        "source_preference": str(signals.get("source_preference") or ""),
        "style_counts": style_counts,
        "posts": posts,
    }


def _publish_task_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "task_id": str(row.get("task_id") or ""),
        "date": _date_text(row.get("trade_date")),
        "title": str(row.get("title") or ""),
        "body": str(row.get("body") or ""),
        "status": str(row.get("status") or ""),
        "quality_score": "" if row.get("quality_score") is None else str(row.get("quality_score")),
    }


def _publish_task_latest_date(connection: Any) -> date | None:
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT trade_date
            FROM {PUBLISH_TASK_TABLE}
            WHERE source_type = %s
              AND trade_date IS NOT NULL
            ORDER BY trade_date DESC
            LIMIT 1
            """,
            ("content_item",),
        )
        row = cursor.fetchone() or {}
    return _parse_date(row.get("trade_date"))


def _publish_task_options(connection: Any) -> dict[str, Any]:
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT DISTINCT trade_date
            FROM {PUBLISH_TASK_TABLE}
            WHERE source_type = %s
              AND trade_date IS NOT NULL
            ORDER BY trade_date DESC
            """,
            ("content_item",),
        )
        dates = [_date_text(row.get("trade_date")) for row in cursor.fetchall()]
    return {"dates": dates}


def _publish_task_filters(params: dict[str, Any], connection: Any) -> dict[str, Any]:
    limit = _parse_int(params.get("limit"), 200) or 200
    trade_date = _parse_date(str(params.get("date") or params.get("trade_date") or "").strip())
    if trade_date is None:
        trade_date = _publish_task_latest_date(connection)
    return {
        "trade_date": trade_date,
        "title": _clean_text(params.get("title") or params.get("q")),
        "limit": min(max(limit, 1), 1000),
    }


def _publish_task_where_sql(filters: dict[str, Any]) -> tuple[str, list[Any]]:
    clauses = ["source_type = %s"]
    args: list[Any] = ["content_item"]
    if filters["trade_date"]:
        clauses.append("trade_date = %s")
        args.append(filters["trade_date"])
    if filters["title"]:
        clauses.append("title LIKE %s")
        args.append(f"%{filters['title']}%")
    return "WHERE " + " AND ".join(clauses), args


def _publish_task_rows(connection: Any, filters: dict[str, Any]) -> list[dict[str, Any]]:
    where_sql, args = _publish_task_where_sql(filters)
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
              task_id,
              trade_date,
              title,
              body,
              status,
              quality_score
            FROM {PUBLISH_TASK_TABLE}
            {where_sql}
            ORDER BY task_id DESC
            LIMIT %s
            """,
            (*args, int(filters["limit"])),
        )
        return [_publish_task_row(row) for row in cursor.fetchall()]


def _publish_task_summary(connection: Any, filters: dict[str, Any]) -> dict[str, int]:
    where_sql, args = _publish_task_where_sql(filters)
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
              COUNT(*) AS total_rows,
              COUNT(DISTINCT trade_date) AS date_count,
              COUNT(DISTINCT title) AS title_count
            FROM {PUBLISH_TASK_TABLE}
            {where_sql}
            """,
            args,
        )
        row = cursor.fetchone() or {}
    return {key: int(value or 0) for key, value in row.items()}


def _publish_tasks_payload(params: dict[str, Any]) -> dict[str, Any]:
    with _connect() as connection:
        filters = _publish_task_filters(params, connection)
        rows = _publish_task_rows(connection, filters)
        return {
            "source": "mysql",
            "table": PUBLISH_TASK_TABLE,
            "filters": {
                **filters,
                "trade_date": _date_text(filters["trade_date"]),
            },
            "options": _publish_task_options(connection),
            "summary": _publish_task_summary(connection, filters),
            "rows": rows,
        }


def _batches_payload(limit: int = 30) -> list[dict[str, Any]]:
    with _connect() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT batch.trade_date, batch.generated_at, batch.run_id, posts.status, COUNT(*) AS post_count
                FROM (
                    SELECT trade_date, generated_at, run_id
                    FROM {POST_TABLE}
                    GROUP BY trade_date, generated_at, run_id
                    ORDER BY trade_date DESC, generated_at DESC, run_id DESC
                    LIMIT %s
                ) AS batch
                JOIN {POST_TABLE} AS posts
                  ON posts.trade_date = batch.trade_date
                 AND posts.generated_at = batch.generated_at
                 AND posts.run_id = batch.run_id
                GROUP BY batch.trade_date, batch.generated_at, batch.run_id, posts.status
                ORDER BY batch.trade_date DESC, batch.generated_at DESC, batch.run_id DESC, posts.status ASC
                """,
                (int(limit),),
            )
            rows = cursor.fetchall()

    batches: list[dict[str, Any]] = []
    batches_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        trade_date = str(row.get("trade_date") or "")
        generated_at = str(row.get("generated_at") or "")
        run_id = str(row.get("run_id") or "")
        status = str(row.get("status") or "")
        post_count = int(row.get("post_count") or 0)
        key = (trade_date, generated_at, run_id)
        batch = batches_by_key.get(key)
        if batch is None:
            batch = {
                "trade_date": trade_date,
                "generated_at": generated_at,
                "run_id": run_id,
                "post_count": 0,
                "status": "",
                "statuses": [],
                "status_counts": {},
            }
            batches_by_key[key] = batch
            batches.append(batch)
        batch["post_count"] += post_count
        if status:
            batch["statuses"].append(status)
            batch["status_counts"][status] = post_count

    for batch in batches:
        statuses = batch["statuses"]
        batch["status"] = statuses[0] if len(statuses) == 1 else "mixed"

    return batches


def _dimensions_payload() -> list[dict[str, Any]]:
    with _connect() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT
                  categories.category_id,
                  categories.name AS category_name,
                  categories.description AS category_description,
                  categories.selection_required,
                  categories.sort_order AS category_sort_order,
                  options.option_id,
                  options.option_key,
                  options.name AS option_name,
                  options.description AS option_description,
                  options.sort_order AS option_sort_order
                FROM {DIMENSION_CATEGORY_TABLE} AS categories
                LEFT JOIN {DIMENSION_OPTION_TABLE} AS options
                  ON options.category_id = categories.category_id
                 AND options.enabled = 1
                WHERE categories.enabled = 1
                ORDER BY categories.sort_order ASC,
                         categories.category_id ASC,
                         options.sort_order ASC,
                         options.option_key ASC
                """
            )
            rows = cursor.fetchall()

    categories: list[dict[str, Any]] = []
    categories_by_id: dict[str, dict[str, Any]] = {}
    for row in rows:
        category_id = str(row.get("category_id") or "")
        if not category_id:
            continue
        category = categories_by_id.get(category_id)
        if category is None:
            category = {
                "category_id": category_id,
                "name": str(row.get("category_name") or ""),
                "description": str(row.get("category_description") or ""),
                "selection_required": bool(row.get("selection_required")),
                "sort_order": int(row.get("category_sort_order") or 0),
                "options": [],
            }
            categories_by_id[category_id] = category
            categories.append(category)

        option_id = str(row.get("option_id") or "")
        if option_id:
            category["options"].append(
                {
                    "option_id": option_id,
                    "category_id": category_id,
                    "option_key": str(row.get("option_key") or ""),
                    "name": str(row.get("option_name") or ""),
                    "description": str(row.get("option_description") or ""),
                    "sort_order": int(row.get("option_sort_order") or 0),
                }
            )
    return categories


def _normalize_post_urls(value: Any) -> list[str]:
    if isinstance(value, str):
        raw_items = re.split(r"[\n\r,]+", value)
    elif isinstance(value, list):
        raw_items = value
    else:
        raw_items = []

    urls: list[str] = []
    seen: set[str] = set()
    for item in raw_items:
        url = str(item or "").strip().strip("'").strip('"')
        if not url or url in seen:
            continue
        seen.add(url)
        urls.append(url)
    return urls


def _task_submission_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": int(row.get("id") or 0),
        "task_type": str(row.get("task_type") or ""),
        "post_url": str(row.get("post_url") or ""),
        "account_name": str(row.get("account_name") or ""),
        "document_id": str(row.get("document_id") or ""),
        "row_index": int(row.get("row_index") or 0),
        "status": str(row.get("status") or ""),
        "attempts": int(row.get("attempts") or 0),
        "updated_at": str(row.get("updated_at") or ""),
    }


def _rerun_post_tasks_payload(urls_value: Any) -> dict[str, Any]:
    urls = _normalize_post_urls(urls_value)
    if not urls:
        raise ValueError("请输入至少一个帖子链接")
    if len(urls) > 500:
        raise ValueError("一次最多处理 500 个帖子链接")

    placeholders = ", ".join(["%s"] * len(urls))
    select_sql = f"""
        SELECT id, task_type, post_url, account_name, document_id, row_index, status, attempts, updated_at
        FROM {TASK_SUBMISSION_TABLE}
        WHERE task_type = %s
          AND post_url IN ({placeholders})
        ORDER BY id ASC
    """

    with _connect() as connection:
        with connection.cursor() as cursor:
            cursor.execute(select_sql, ("detail", *urls))
            matched_rows = cursor.fetchall()

        matched_ids = [int(row["id"]) for row in matched_rows if row.get("id")]
        if matched_ids:
            id_placeholders = ", ".join(["%s"] * len(matched_ids))
            with connection.cursor() as cursor:
                cursor.execute(
                    f"""
                    UPDATE {TASK_SUBMISSION_TABLE}
                    SET status = %s,
                        attempts = %s
                    WHERE task_type = %s
                      AND id IN ({id_placeholders})
                    """,
                    ("pending", 1, "detail", *matched_ids),
                )
                updated_count = int(cursor.rowcount or 0)
            connection.commit()
        else:
            updated_count = 0

        if matched_ids:
            id_placeholders = ", ".join(["%s"] * len(matched_ids))
            with connection.cursor() as cursor:
                cursor.execute(
                    f"""
                    SELECT id, task_type, post_url, account_name, document_id, row_index, status, attempts, updated_at
                    FROM {TASK_SUBMISSION_TABLE}
                    WHERE id IN ({id_placeholders})
                    ORDER BY id ASC
                    """,
                    tuple(matched_ids),
                )
                updated_rows = cursor.fetchall()
        else:
            updated_rows = []

    matched_urls = {str(row.get("post_url") or "") for row in matched_rows}
    return {
        "requested_count": len(urls),
        "matched_count": len(matched_rows),
        "updated_count": updated_count,
        "unmatched_urls": [url for url in urls if url not in matched_urls],
        "rows": [_task_submission_row(row) for row in updated_rows],
    }


def _settlement_autofill_payload(post_url: str, request: FastAPIRequest | None = None) -> dict[str, Any]:
    post_url = post_url.strip()
    if not post_url:
        raise ValueError("请输入帖子链接")

    payload: dict[str, Any] = {
        "link": post_url,
        "date": "",
        "product": "",
        "ipName": "",
        "fansCount": "",
        "articleTitle": "",
        "screenshot": "",
        "readCount": "",
        "commentCount": "",
        "likeCount": "",
        "sources": [],
    }

    with _connect() as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT ts.account_name, ts.post_time, ts.source_locator_json,
                       te.result_json, te.metrics_json, te.screenshot_path, te.finished_at
                FROM crawler_app.task_submissions AS ts
                LEFT JOIN crawler_app.task_executions AS te
                  ON te.id = ts.latest_execution_id
                WHERE ts.post_url = %s
                ORDER BY ts.updated_at DESC
                LIMIT 1
                """,
                (post_url,),
            )
            task_row = cursor.fetchone() or {}

            cursor.execute(
                """
                SELECT ip_name, product_name, first_seen_date, latest_seen_date, source_json
                FROM crawler_app.article_detail_targets
                WHERE article_url = %s
                ORDER BY updated_at DESC
                LIMIT 1
                """,
                (post_url,),
            )
            target_row = cursor.fetchone() or {}

            cursor.execute(
                """
                SELECT article_title, read_count, comment_count, like_count,
                       metrics_json, screenshot_path, crawled_at
                FROM crawler_app.article_detail_runs
                WHERE article_url = %s
                ORDER BY crawled_at DESC
                LIMIT 1
                """,
                (post_url,),
            )
            detail_row = cursor.fetchone() or {}

        ip_name = str(_first_present(target_row.get("ip_name"), task_row.get("account_name")) or "")
        if ip_name:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT fans_count, snapshot_date
                    FROM crawler_app.kol_daily_snapshots
                    WHERE kol_name = %s
                      AND fans_count IS NOT NULL
                    ORDER BY snapshot_date DESC
                    LIMIT 1
                    """,
                    (ip_name,),
                )
                snapshot_row = cursor.fetchone() or {}

                cursor.execute(
                    """
                    SELECT fans_count, metric_date
                    FROM crawler_app.kol_daily_metrics
                    WHERE kol_name = %s
                      AND fans_count IS NOT NULL
                    ORDER BY metric_date DESC
                    LIMIT 1
                    """,
                    (ip_name,),
                )
                metric_row = cursor.fetchone() or {}
        else:
            snapshot_row = {}
            metric_row = {}

    task_result = _json_value(task_row.get("result_json"), {})
    task_metrics = _json_value(task_row.get("metrics_json"), {})
    target_source = _json_value(target_row.get("source_json"), {})
    detail_metrics = _json_value(detail_row.get("metrics_json"), {})

    payload["date"] = _date_text(
        _first_present(
            target_row.get("latest_seen_date"),
            target_row.get("first_seen_date"),
            detail_row.get("crawled_at"),
            task_row.get("post_time"),
            task_row.get("finished_at"),
        )
    )
    payload["product"] = str(_first_present(target_row.get("product_name"), _deep_find_value(target_source, {"product", "product_name"})) or "")
    payload["ipName"] = ip_name
    payload["fansCount"] = str(_first_present(snapshot_row.get("fans_count"), metric_row.get("fans_count"), _deep_find_value(task_metrics, {"fans_count", "fansCount"}), _deep_find_value(task_result, {"fans_count", "fansCount"})) or "")
    payload["articleTitle"] = str(_first_present(detail_row.get("article_title"), _deep_find_value(detail_metrics, {"article_title", "articleTitle", "title"}), _deep_find_value(task_metrics, {"article_title", "articleTitle", "title"}), _deep_find_value(task_result, {"article_title", "articleTitle", "title"})) or "")
    payload["screenshot"] = _externalize_local_url(_first_present(detail_row.get("screenshot_path"), task_row.get("screenshot_path"), _deep_find_value(detail_metrics, {"screenshot", "screenshot_path", "screenshotPath"}), _deep_find_value(task_result, {"screenshot", "screenshot_path", "screenshotPath"})), request)
    payload["readCount"] = str(_first_present(detail_row.get("read_count"), _deep_find_value(detail_metrics, {"read_count", "readCount", "reads"}), _deep_find_value(task_metrics, {"read_count", "readCount", "reads"}), _deep_find_value(task_result, {"read_count", "readCount", "reads"})) or "")
    payload["commentCount"] = str(_first_present(detail_row.get("comment_count"), _deep_find_value(detail_metrics, {"comment_count", "commentCount", "comments"}), _deep_find_value(task_metrics, {"comment_count", "commentCount", "comments"}), _deep_find_value(task_result, {"comment_count", "commentCount", "comments"})) or "")
    payload["likeCount"] = str(_first_present(detail_row.get("like_count"), _deep_find_value(detail_metrics, {"like_count", "likeCount", "likes"}), _deep_find_value(task_metrics, {"like_count", "likeCount", "likes"}), _deep_find_value(task_result, {"like_count", "likeCount", "likes"})) or "")

    if task_row:
        payload["sources"].append("task_submissions")
    if target_row:
        payload["sources"].append("article_detail_targets")
    if detail_row:
        payload["sources"].append("article_detail_runs")
    if snapshot_row:
        payload["sources"].append("kol_daily_snapshots")
    if metric_row:
        payload["sources"].append("kol_daily_metrics")
    return payload


SETTLEMENT_IMPORT_FIELDS = [
    "date",
    "partner",
    "deliveryPlatform",
    "product",
    "productCode",
    "ipName",
    "fansCount",
    "articleType",
    "fee",
    "creatorFee",
    "kolType",
    "buyAmount",
    "link",
    "articleTitle",
    "screenshot",
    "readCount",
    "commentCount",
    "likeCount",
    "partnerPaymentStatus",
    "creatorSettlementStatus",
    "notes",
]

SETTLEMENT_EXPORT_LABELS = {
    "date": "日期",
    "partner": "合作方",
    "deliveryPlatform": "投放平台",
    "product": "产品",
    "productCode": "代码",
    "ipName": "IP名称",
    "fansCount": "粉丝数",
    "articleType": "文章类型",
    "fee": "费用",
    "creatorFee": "创作者费用",
    "kolType": "大V类型",
    "buyAmount": "买入金额",
    "link": "链接",
    "articleTitle": "文章标题",
    "screenshot": "截图",
    "readCount": "阅读量",
    "commentCount": "评论",
    "likeCount": "点赞",
    "partnerPaymentStatus": "打款进度-合作方",
    "creatorSettlementStatus": "结算进度-创作者",
    "notes": "备注",
}

SETTLEMENT_EXPORT_NUMERIC_FIELDS = {"fansCount", "fee", "creatorFee", "buyAmount", "readCount", "commentCount", "likeCount"}
SETTLEMENT_EXPORT_MONEY_FIELDS = {"fee", "creatorFee", "buyAmount"}

SETTLEMENT_DB_COLUMNS = [
    "settlement_date",
    "partner",
    "delivery_platform",
    "product_name",
    "product_code",
    "ip_name",
    "fans_count",
    "article_type",
    "fee",
    "creator_fee",
    "kol_type",
    "buy_amount",
    "post_url",
    "post_url_hash",
    "article_title",
    "screenshot_url",
    "read_count",
    "comment_count",
    "like_count",
    "partner_payment_status",
    "creator_settlement_status",
    "notes",
    "source_payload_json",
]
SETTLEMENT_NUMERIC_DB_COLUMNS = {"fans_count", "fee", "creator_fee", "buy_amount", "read_count", "comment_count", "like_count"}
SETTLEMENT_AUTOFILL_DB_COLUMNS = {"fans_count", "article_title", "screenshot_url", "read_count", "comment_count", "like_count"}
SETTLEMENT_ZERO_AS_EMPTY_NUMERIC_COLUMNS = {"fans_count"}
SETTLEMENT_TEXT_PLACEHOLDERS = ("0", "机器识别", "自动识别", "程序识别", "待识别", "最好程序能帮填写")
SETTLEMENT_REQUIRED_TEXT_DB_COLUMNS = {
    "partner",
    "delivery_platform",
    "product_name",
    "ip_name",
    "article_type",
    "post_url",
    "post_url_hash",
    "screenshot_url",
    "partner_payment_status",
    "creator_settlement_status",
    "product_code",
}


def _settlement_product_code_from_text(value: Any) -> str:
    match = re.search(r"(?<!\d)(\d{6})(?!\d)", str(value or ""))
    return match.group(1) if match else ""


def _clean_settlement_product_code(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    code = _settlement_product_code_from_text(text)
    if code:
        return code
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        number = str(int(float(text)))
        if 1 <= len(number) <= 6:
            return number.zfill(6)
    return text


def _settlement_product_name_key(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"(?<!\d)\d{6}(?!\d)", "", text)
    text = re.sub(r"[\s（）()【】\[\]·,，;；:：-]+", "", text)
    return text


def _settlement_product_code_map(cursor: Any) -> dict[str, str]:
    cursor.execute(
        f"""
        SELECT product_name, product_code
        FROM {SETTLEMENT_TABLE}
        WHERE product_name <> ''
          AND product_code IS NOT NULL
          AND product_code <> ''
        """
    )
    candidates: dict[str, set[str]] = {}
    for row in cursor.fetchall():
        key = _settlement_product_name_key(row.get("product_name"))
        code = _clean_settlement_product_code(row.get("product_code"))
        if key and code:
            candidates.setdefault(key, set()).add(code)
    return {key: next(iter(codes)) for key, codes in candidates.items() if len(codes) == 1}


def _backfill_settlement_product_codes(cursor: Any) -> None:
    def update_without_identity_conflict(rows: list[dict[str, Any]], code_getter: Callable[[dict[str, Any]], str]) -> None:
        updates = []
        for row in rows:
            code = _clean_settlement_product_code(code_getter(row))
            row_id = row.get("id")
            if not code or not row_id:
                continue
            cursor.execute(
                f"""
                SELECT id
                FROM {SETTLEMENT_TABLE}
                WHERE id <> %s
                  AND settlement_date = %s
                  AND ip_name = %s
                  AND article_type = %s
                  AND product_code = %s
                LIMIT 1
                """,
                (row_id, row.get("settlement_date"), row.get("ip_name"), row.get("article_type"), code),
            )
            if cursor.fetchone():
                continue
            updates.append((code, row_id))
        if not updates:
            return
        cursor.executemany(f"UPDATE {SETTLEMENT_TABLE} SET product_code = %s WHERE id = %s", updates)

    cursor.execute(f"SELECT id, settlement_date, ip_name, product_name, product_code, article_type FROM {SETTLEMENT_TABLE} WHERE (product_code IS NULL OR product_code = '') AND product_name <> ''")
    update_without_identity_conflict(cursor.fetchall(), lambda row: _settlement_product_code_from_text(row.get("product_name")))

    known_codes = _settlement_product_code_map(cursor)
    if not known_codes:
        return
    cursor.execute(f"SELECT id, settlement_date, ip_name, product_name, product_code, article_type FROM {SETTLEMENT_TABLE} WHERE (product_code IS NULL OR product_code = '') AND product_name <> ''")
    update_without_identity_conflict(cursor.fetchall(), lambda row: known_codes.get(_settlement_product_name_key(row.get("product_name")), ""))
    cursor.execute(f"UPDATE {SETTLEMENT_TABLE} SET product_code = NULL WHERE product_code = ''")


def _settlement_value_present(column: str, value: Any) -> bool:
    if value is None:
        return False
    if column in SETTLEMENT_NUMERIC_DB_COLUMNS:
        return True
    return _clean_text(value) != ""


def _settlement_collected_score(row: dict[str, Any]) -> int:
    return sum(1 for column in SETTLEMENT_AUTOFILL_DB_COLUMNS if _settlement_value_present(column, row.get(column)))


def _merged_settlement_row(rows: list[dict[str, Any]]) -> dict[str, Any]:
    collected_rows = sorted(rows, key=lambda row: (-_settlement_collected_score(row), int(row.get("id") or 0)))
    business_rows = sorted(rows, key=lambda row: int(row.get("id") or 0), reverse=True)
    merged: dict[str, Any] = {}
    for column in SETTLEMENT_DB_COLUMNS:
        candidates = collected_rows if column in SETTLEMENT_AUTOFILL_DB_COLUMNS else business_rows
        merged[column] = None
        for row in candidates:
            value = row.get(column)
            if _settlement_value_present(column, value):
                merged[column] = value
                break
        if merged[column] is None and column in SETTLEMENT_REQUIRED_TEXT_DB_COLUMNS:
            merged[column] = ""
    merged["post_url_hash"] = _settlement_url_hash(_clean_text(merged.get("post_url")))
    return merged


def _merge_settlement_duplicate_identities(cursor: Any) -> int:
    cursor.execute(
        f"""
        SELECT GROUP_CONCAT(id ORDER BY id) AS ids
        FROM {SETTLEMENT_TABLE}
        WHERE product_code IS NOT NULL
          AND product_code <> ''
        GROUP BY settlement_date, ip_name, product_code, article_type
        HAVING COUNT(*) > 1
        """
    )
    groups = cursor.fetchall()
    merged_count = 0
    update_assignments = ", ".join(f"{column} = %s" for column in SETTLEMENT_DB_COLUMNS)
    for group in groups:
        ids = [int(value) for value in str(group.get("ids") or "").split(",") if value.strip().isdigit()]
        if len(ids) < 2:
            continue
        placeholders = ", ".join(["%s"] * len(ids))
        cursor.execute(
            f"""
            SELECT id, {", ".join(SETTLEMENT_DB_COLUMNS)}
            FROM {SETTLEMENT_TABLE}
            WHERE id IN ({placeholders})
            """,
            tuple(ids),
        )
        rows = cursor.fetchall()
        if len(rows) < 2:
            continue
        keeper = sorted(rows, key=lambda row: (-_settlement_collected_score(row), int(row.get("id") or 0)))[0]
        keeper_id = int(keeper.get("id") or 0)
        duplicate_ids = [int(row.get("id") or 0) for row in rows if int(row.get("id") or 0) != keeper_id]
        if not keeper_id or not duplicate_ids:
            continue
        merged = _merged_settlement_row(rows)
        delete_placeholders = ", ".join(["%s"] * len(duplicate_ids))
        cursor.execute(f"DELETE FROM {SETTLEMENT_TABLE} WHERE id IN ({delete_placeholders})", tuple(duplicate_ids))
        cursor.execute(
            f"""
            UPDATE {SETTLEMENT_TABLE}
            SET {update_assignments}
            WHERE id = %s
            """,
            (*tuple(merged[column] for column in SETTLEMENT_DB_COLUMNS), keeper_id),
        )
        merged_count += len(duplicate_ids)
    return merged_count


def _settlement_index_columns(cursor: Any, index_name: str) -> list[str]:
    cursor.execute(f"SHOW INDEX FROM {SETTLEMENT_TABLE} WHERE Key_name = %s", (index_name,))
    return [str(row.get("Column_name") or "") for row in cursor.fetchall()]


def _settlement_new_identity_duplicate_count(cursor: Any) -> int:
    cursor.execute(
        f"""
        SELECT COUNT(*) AS duplicate_groups
        FROM (
          SELECT settlement_date, ip_name, product_code, article_type
          FROM {SETTLEMENT_TABLE}
          WHERE product_code IS NOT NULL
            AND product_code <> ''
          GROUP BY settlement_date, ip_name, product_code, article_type
          HAVING COUNT(*) > 1
        ) AS duplicate_keys
        """
    )
    return int((cursor.fetchone() or {}).get("duplicate_groups") or 0)


def _ensure_settlement_table(connection: Any) -> None:
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            CREATE TABLE IF NOT EXISTS {SETTLEMENT_TABLE} (
              id BIGINT UNSIGNED NOT NULL AUTO_INCREMENT,
              settlement_date DATE NOT NULL,
              partner VARCHAR(255) NOT NULL DEFAULT '',
              delivery_platform VARCHAR(255) NOT NULL DEFAULT '',
              product_name VARCHAR(255) NOT NULL DEFAULT '',
              product_code VARCHAR(64) NULL DEFAULT NULL,
              ip_name VARCHAR(255) NOT NULL DEFAULT '',
              fans_count BIGINT NULL,
              article_type VARCHAR(32) NOT NULL DEFAULT '',
              fee DECIMAL(14,2) NULL,
              creator_fee DECIMAL(14,2) NULL,
              kol_type VARCHAR(32) NOT NULL DEFAULT '',
              buy_amount DECIMAL(14,2) NULL,
              post_url VARCHAR(1000) NOT NULL,
              post_url_hash CHAR(64) NOT NULL,
              article_title TEXT NULL,
              screenshot_url VARCHAR(1000) NOT NULL DEFAULT '',
              read_count BIGINT NULL,
              comment_count BIGINT NULL,
              like_count BIGINT NULL,
              partner_payment_status VARCHAR(64) NOT NULL DEFAULT '',
              creator_settlement_status VARCHAR(64) NOT NULL DEFAULT '',
              notes TEXT NULL,
              source_payload_json LONGTEXT NULL,
              created_at DATETIME NULL DEFAULT CURRENT_TIMESTAMP,
              updated_at DATETIME NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
              PRIMARY KEY (id),
              UNIQUE KEY uq_settlement_identity (settlement_date, ip_name, product_code, article_type),
              KEY idx_settlement_date (settlement_date),
              KEY idx_ip_name (ip_name),
              KEY idx_product_code (product_code),
              KEY idx_delivery_platform (delivery_platform),
              KEY idx_settlement_post_url_hash (post_url_hash),
              KEY idx_partner (partner)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
            """
        )
        cursor.execute(f"SHOW COLUMNS FROM {SETTLEMENT_TABLE} LIKE 'delivery_platform'")
        if not cursor.fetchone():
            cursor.execute(
                f"""
                ALTER TABLE {SETTLEMENT_TABLE}
                ADD COLUMN delivery_platform VARCHAR(255) NOT NULL DEFAULT '' AFTER partner
                """
            )
        cursor.execute(f"SHOW COLUMNS FROM {SETTLEMENT_TABLE} LIKE 'product_code'")
        if not cursor.fetchone():
            cursor.execute(
                f"""
                ALTER TABLE {SETTLEMENT_TABLE}
                ADD COLUMN product_code VARCHAR(64) NULL DEFAULT NULL AFTER product_name
                """
            )
        else:
            cursor.execute(f"ALTER TABLE {SETTLEMENT_TABLE} MODIFY product_code VARCHAR(64) NULL DEFAULT NULL")
        _backfill_settlement_product_codes(cursor)
        _merge_settlement_duplicate_identities(cursor)
        cursor.execute(f"SHOW INDEX FROM {SETTLEMENT_TABLE} WHERE Key_name = 'idx_delivery_platform'")
        if not cursor.fetchone():
            cursor.execute(f"ALTER TABLE {SETTLEMENT_TABLE} ADD KEY idx_delivery_platform (delivery_platform)")
        cursor.execute(f"SHOW INDEX FROM {SETTLEMENT_TABLE} WHERE Key_name = 'idx_product_code'")
        if not cursor.fetchone():
            cursor.execute(f"ALTER TABLE {SETTLEMENT_TABLE} ADD KEY idx_product_code (product_code)")
        identity_columns = _settlement_index_columns(cursor, "uq_settlement_identity")
        expected_identity_columns = ["settlement_date", "ip_name", "product_code", "article_type"]
        if identity_columns != expected_identity_columns and _settlement_new_identity_duplicate_count(cursor) == 0:
            if identity_columns:
                cursor.execute(f"ALTER TABLE {SETTLEMENT_TABLE} DROP INDEX uq_settlement_identity")
            cursor.execute(
                f"""
                ALTER TABLE {SETTLEMENT_TABLE}
                ADD UNIQUE KEY uq_settlement_identity (settlement_date, ip_name, product_code, article_type)
                """
            )
        cursor.execute(f"SHOW INDEX FROM {SETTLEMENT_TABLE} WHERE Key_name = 'idx_settlement_post_url_hash'")
        if not cursor.fetchone():
            cursor.execute(f"ALTER TABLE {SETTLEMENT_TABLE} ADD KEY idx_settlement_post_url_hash (post_url_hash)")
        cursor.execute(f"SHOW INDEX FROM {SETTLEMENT_TABLE} WHERE Key_name = 'uq_settlement_date_url'")
        if cursor.fetchone():
            cursor.execute(f"ALTER TABLE {SETTLEMENT_TABLE} DROP INDEX uq_settlement_date_url")


def _settlement_url_hash(post_url: str) -> str:
    return hashlib.sha256(post_url.strip().encode("utf-8")).hexdigest()


def _settlement_import_update_sql(columns: list[str]) -> str:
    assignments: list[str] = []
    text_placeholder_sql = " OR ".join(
        f"VALUES({{column}}) LIKE CONCAT(CHAR(37), '{placeholder}', CHAR(37))"
        for placeholder in SETTLEMENT_TEXT_PLACEHOLDERS
        if placeholder != "0"
    )
    for column in columns:
        if column in SETTLEMENT_AUTOFILL_DB_COLUMNS and column in SETTLEMENT_ZERO_AS_EMPTY_NUMERIC_COLUMNS:
            assignments.append(f"{column} = CASE WHEN VALUES({column}) IS NULL OR VALUES({column}) = 0 THEN {column} ELSE VALUES({column}) END")
        elif column in SETTLEMENT_AUTOFILL_DB_COLUMNS and column in SETTLEMENT_NUMERIC_DB_COLUMNS:
            assignments.append(f"{column} = COALESCE(VALUES({column}), {column})")
        elif column in SETTLEMENT_AUTOFILL_DB_COLUMNS:
            placeholder_sql = text_placeholder_sql.format(column=column)
            assignments.append(f"{column} = CASE WHEN VALUES({column}) IS NULL OR VALUES({column}) = '' OR VALUES({column}) = '0' OR {placeholder_sql} THEN {column} ELSE VALUES({column}) END")
        elif column in SETTLEMENT_NUMERIC_DB_COLUMNS:
            assignments.append(f"{column} = COALESCE(VALUES({column}), {column})")
        else:
            assignments.append(f"{column} = CASE WHEN VALUES({column}) IS NULL OR VALUES({column}) = '' THEN {column} ELSE VALUES({column}) END")
    return ", ".join(assignments)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _first_non_empty_value(*values: Any) -> Any:
    for value in values:
        if value is not None and _clean_text(value) != "":
            return value
    return None


def _is_zero_placeholder(value: Any) -> bool:
    return re.fullmatch(r"0+(?:\.0+)?", _clean_text(value)) is not None


def _is_settlement_text_placeholder(value: Any) -> bool:
    text = _clean_text(value)
    if not text:
        return False
    return _is_zero_placeholder(text) or any(placeholder != "0" and placeholder in text for placeholder in SETTLEMENT_TEXT_PLACEHOLDERS)


def _settlement_identity_key(row: dict[str, Any]) -> tuple[str, str, str, str]:
    return (
        _date_text(row.get("settlement_date")),
        _clean_text(row.get("ip_name")),
        _clean_text(row.get("product_code")),
        _clean_text(row.get("article_type")),
    )


def _clean_date(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")
    text = _clean_text(value).replace("/", "-")
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        serial = int(float(text))
        if 20000 <= serial <= 80000:
            return (datetime(1899, 12, 30) + timedelta(days=serial)).strftime("%Y-%m-%d")
    match = re.search(r"\d{4}-\d{1,2}-\d{1,2}", text)
    if not match:
        raise ValueError("导入行缺少有效日期，格式应为 YYYY-MM-DD")
    year, month, day = match.group(0).split("-")
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def _clean_int(value: Any) -> int | None:
    text = _clean_text(value).replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+", text)
    return int(match.group(0)) if match else None


def _clean_decimal(value: Any) -> str | None:
    text = _clean_text(value).replace(",", "").replace("￥", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return match.group(0) if match else None


def _compact_header(value: Any) -> str:
    return re.sub(r"[\s\r\n（）()]+", "", _clean_text(value))


def _settlement_field_from_header(header: Any) -> str:
    text = _compact_header(header)
    if not text:
        return ""
    if "创作者费用" in text:
        return "creatorFee"
    if "打款进度" in text and "合作方" in text:
        return "partnerPaymentStatus"
    if "结算进度" in text and "创作者" in text:
        return "creatorSettlementStatus"
    checks = [
        ("日期", "date"),
        ("合作方", "partner"),
        ("投放平台", "deliveryPlatform"),
        ("基金代码", "productCode"),
        ("产品代码", "productCode"),
        ("代码", "productCode"),
        ("产品", "product"),
        ("IP名称", "ipName"),
        ("粉丝数", "fansCount"),
        ("文章类型", "articleType"),
        ("费用", "fee"),
        ("大V类型", "kolType"),
        ("买入金额", "buyAmount"),
        ("链接", "link"),
        ("文章标题", "articleTitle"),
        ("截图", "screenshot"),
        ("阅读量", "readCount"),
        ("评论", "commentCount"),
        ("点赞", "likeCount"),
        ("备注", "notes"),
    ]
    for key, field in checks:
        if key in text:
            return field
    return ""


def _rows_from_grid(grid: list[list[Any]]) -> list[dict[str, Any]]:
    non_empty_rows = [row for row in grid if any(_clean_text(cell) for cell in row)]
    if len(non_empty_rows) < 2:
        return []
    header_index = 0
    for index, row in enumerate(non_empty_rows[:10]):
        fields = [_settlement_field_from_header(cell) for cell in row]
        if "date" in fields and "link" in fields:
            header_index = index
            break
    headers = [_settlement_field_from_header(cell) for cell in non_empty_rows[header_index]]
    rows: list[dict[str, Any]] = []
    for raw_row in non_empty_rows[header_index + 1 :]:
        item: dict[str, Any] = {}
        for index, field in enumerate(headers):
            if field:
                item[field] = raw_row[index] if index < len(raw_row) else ""
        if item.get("date") or item.get("link"):
            rows.append(item)
    return rows


def _parse_delimited_settlement_file(content: bytes, filename: str) -> list[dict[str, Any]]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    delimiter = "\t" if "\t" in sample else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return _rows_from_grid([list(row) for row in reader])


def _xlsx_col_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha())
    index = 0
    for char in letters:
        index = index * 26 + ord(char.upper()) - 64
    return max(index - 1, 0)


def _xlsx_shared_strings(zip_file: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in zip_file.namelist():
        return []
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    root = ET.fromstring(zip_file.read("xl/sharedStrings.xml"))
    strings: list[str] = []
    for item in root.findall("a:si", ns):
        strings.append("".join(text.text or "" for text in item.findall(".//a:t", ns)))
    return strings


def _parse_xlsx_settlement_file(content: bytes) -> list[dict[str, Any]]:
    ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    with zipfile.ZipFile(io.BytesIO(content)) as zip_file:
        workbook = ET.fromstring(zip_file.read("xl/workbook.xml"))
        sheets = [
            {
                "name": sheet.attrib.get("name", ""),
                "rid": sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", ""),
            }
            for sheet in workbook.findall("a:sheets/a:sheet", ns)
        ]
        rels = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))
        rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
        selected = next((sheet for sheet in sheets if sheet["name"] == "总表"), None)
        if selected is None:
            selected = next((sheet for sheet in sheets if "总" in sheet["name"]), None)
        if selected is None:
            selected = sheets[1] if len(sheets) > 1 else sheets[0]
        sheet_path = "xl/" + rel_map[selected["rid"]].lstrip("/")
        shared_strings = _xlsx_shared_strings(zip_file)
        root = ET.fromstring(zip_file.read(sheet_path))

        def cell_value(cell: ET.Element) -> str:
            cell_type = cell.attrib.get("t")
            if cell_type == "inlineStr":
                return "".join(text.text or "" for text in cell.findall(".//a:t", ns))
            value = cell.find("a:v", ns)
            if value is None:
                return ""
            raw = value.text or ""
            if cell_type == "s":
                index = int(raw) if raw.isdigit() else -1
                return shared_strings[index] if 0 <= index < len(shared_strings) else raw
            return raw

        grid: list[list[str]] = []
        for row in root.findall("a:sheetData/a:row", ns):
            values: list[str] = []
            for cell in row.findall("a:c", ns):
                index = _xlsx_col_index(cell.attrib.get("r", "A1"))
                while len(values) <= index:
                    values.append("")
                values[index] = cell_value(cell)
            grid.append(values)
    return _rows_from_grid(grid)


def _parse_settlement_import_file(content: bytes, filename: str) -> list[dict[str, Any]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".xlsx":
        return _parse_xlsx_settlement_file(content)
    if suffix in {".csv", ".tsv", ".txt"}:
        return _parse_delimited_settlement_file(content, filename)
    raise ValueError("仅支持 .xlsx / .csv / .tsv / .txt 文件")


def _settlement_api_row(row: dict[str, Any], request: FastAPIRequest | None = None) -> dict[str, Any]:
    return {
        "id": str(row.get("id") or ""),
        "date": _date_text(row.get("settlement_date")),
        "partner": str(row.get("partner") or ""),
        "deliveryPlatform": str(row.get("delivery_platform") or ""),
        "product": str(row.get("product_name") or ""),
        "productCode": str(row.get("product_code") or ""),
        "ipName": str(row.get("ip_name") or ""),
        "fansCount": "" if row.get("fans_count") is None else str(row.get("fans_count")),
        "articleType": str(row.get("article_type") or ""),
        "fee": "" if row.get("fee") is None else str(row.get("fee")),
        "creatorFee": "" if row.get("creator_fee") is None else str(row.get("creator_fee")),
        "kolType": str(row.get("kol_type") or ""),
        "buyAmount": "" if row.get("buy_amount") is None else str(row.get("buy_amount")),
        "link": str(row.get("post_url") or ""),
        "articleTitle": str(row.get("article_title") or ""),
        "screenshot": _externalize_local_url(row.get("screenshot_url"), request),
        "readCount": "" if row.get("read_count") is None else str(row.get("read_count")),
        "commentCount": "" if row.get("comment_count") is None else str(row.get("comment_count")),
        "likeCount": "" if row.get("like_count") is None else str(row.get("like_count")),
        "partnerPaymentStatus": str(row.get("partner_payment_status") or ""),
        "creatorSettlementStatus": str(row.get("creator_settlement_status") or ""),
        "notes": str(row.get("notes") or ""),
    }


def _normalize_settlement_row(raw: dict[str, Any]) -> dict[str, Any]:
    post_url = _clean_text(_first_non_empty_value(raw.get("link"), raw.get("post_url"), raw.get("链接")))
    settlement_date = _clean_date(_first_non_empty_value(raw.get("date"), raw.get("settlement_date"), raw.get("日期")))
    normalized = {
        "settlement_date": settlement_date,
        "partner": _clean_text(_first_non_empty_value(raw.get("partner"), raw.get("合作方"))),
        "delivery_platform": _clean_text(_first_non_empty_value(raw.get("deliveryPlatform"), raw.get("delivery_platform"), raw.get("投放平台"))),
        "product_name": _clean_text(_first_non_empty_value(raw.get("product"), raw.get("product_name"), raw.get("产品"))),
        "product_code": _clean_settlement_product_code(
            _first_non_empty_value(raw.get("productCode"), raw.get("product_code"), raw.get("code"), raw.get("fundCode"), raw.get("fund_code"), raw.get("代码"), raw.get("基金代码"), raw.get("产品代码"))
        ),
        "ip_name": _clean_text(_first_non_empty_value(raw.get("ipName"), raw.get("ip_name"), raw.get("IP名称"))),
        "fans_count": _clean_int(_first_non_empty_value(raw.get("fansCount"), raw.get("fans_count"), raw.get("粉丝数"))),
        "article_type": _clean_text(_first_non_empty_value(raw.get("articleType"), raw.get("article_type"), raw.get("文章类型"))),
        "fee": _clean_decimal(_first_non_empty_value(raw.get("fee"), raw.get("费用"))),
        "creator_fee": _clean_decimal(_first_non_empty_value(raw.get("creatorFee"), raw.get("creator_fee"), raw.get("创作者费用"))),
        "kol_type": _clean_text(_first_non_empty_value(raw.get("kolType"), raw.get("kol_type"), raw.get("大V类型"))),
        "buy_amount": _clean_decimal(_first_non_empty_value(raw.get("buyAmount"), raw.get("buy_amount"), raw.get("买入金额"))),
        "post_url": post_url,
        "post_url_hash": _settlement_url_hash(post_url),
        "article_title": _clean_text(_first_non_empty_value(raw.get("articleTitle"), raw.get("article_title"), raw.get("文章标题"))),
        "screenshot_url": _clean_text(_first_non_empty_value(raw.get("screenshot"), raw.get("screenshot_url"), raw.get("截图"))),
        "read_count": _clean_int(_first_non_empty_value(raw.get("readCount"), raw.get("read_count"), raw.get("阅读量"))),
        "comment_count": _clean_int(_first_non_empty_value(raw.get("commentCount"), raw.get("comment_count"), raw.get("评论"))),
        "like_count": _clean_int(_first_non_empty_value(raw.get("likeCount"), raw.get("like_count"), raw.get("点赞"))),
        "partner_payment_status": _clean_text(_first_non_empty_value(raw.get("partnerPaymentStatus"), raw.get("partner_payment_status"), raw.get("打款进度-合作方"))),
        "creator_settlement_status": _clean_text(_first_non_empty_value(raw.get("creatorSettlementStatus"), raw.get("creator_settlement_status"), raw.get("结算进度-创作者"))),
        "notes": _clean_text(_first_non_empty_value(raw.get("notes"), raw.get("备注"))),
        "source_payload_json": json.dumps(raw, ensure_ascii=False, default=str),
    }
    if not normalized["product_code"]:
        normalized["product_code"] = _settlement_product_code_from_text(normalized["product_name"])
    if not normalized["product_code"]:
        normalized["product_code"] = None
    for metric_column in ("fans_count",):
        if normalized[metric_column] is not None and normalized[metric_column] <= 0:
            normalized[metric_column] = None
    for metric_column in ("read_count", "comment_count", "like_count"):
        if normalized[metric_column] is not None and normalized[metric_column] < 0:
            normalized[metric_column] = None
    for text_column in ("article_title", "screenshot_url"):
        if _is_settlement_text_placeholder(normalized[text_column]):
            normalized[text_column] = ""
    return normalized


def _settlements_payload(limit: int = 500, request: FastAPIRequest | None = None) -> list[dict[str, Any]]:
    with _connect() as connection:
        _ensure_settlement_table(connection)
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT *
                FROM {SETTLEMENT_TABLE}
                ORDER BY settlement_date DESC, id DESC
                LIMIT %s
                """,
                (int(limit),),
            )
            rows = cursor.fetchall()
    return [_settlement_api_row(row, request) for row in rows]


SETTLEMENT_EDITABLE_ENGAGEMENT_FIELDS = {
    "readCount": "read_count",
    "commentCount": "comment_count",
    "likeCount": "like_count",
}


def _settlement_edit_metric_value(value: Any) -> int | None:
    if value is None or _clean_text(value) == "":
        return None
    number = _clean_int(value)
    if number is None or number < 0:
        raise ValueError("阅读量、评论和点赞必须是非负整数，或留空")
    return number


def _update_settlement_engagement_payload(payload: dict[str, Any], request: FastAPIRequest | None = None) -> dict[str, Any]:
    if not isinstance(payload, dict):
        raise ValueError("缺少更新内容")
    row_id = _parse_int(payload.get("id"))
    if row_id is None or row_id <= 0:
        raise ValueError("缺少有效记录 ID")

    field = _clean_text(payload.get("field"))
    if field not in SETTLEMENT_EDITABLE_ENGAGEMENT_FIELDS:
        raise ValueError("只能修改阅读量、评论或点赞")
    column = SETTLEMENT_EDITABLE_ENGAGEMENT_FIELDS[field]
    value = _settlement_edit_metric_value(payload.get("value"))

    with _connect() as connection:
        _ensure_settlement_table(connection)
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                UPDATE {SETTLEMENT_TABLE}
                SET {column} = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                """,
                (value, row_id),
            )
            cursor.execute(f"SELECT id FROM {SETTLEMENT_TABLE} WHERE id = %s", (row_id,))
            if not cursor.fetchone():
                raise ValueError("记录不存在")
        connection.commit()

    return {
        "updated_id": str(row_id),
        "field": field,
        "value": "" if value is None else str(value),
        "rows": _settlements_payload(request=request),
    }


def _delete_settlement_payload(row_id_value: Any, request: FastAPIRequest | None = None) -> dict[str, Any]:
    row_id = _parse_int(row_id_value)
    if row_id is None or row_id <= 0:
        raise ValueError("缺少有效记录 ID")

    with _connect() as connection:
        _ensure_settlement_table(connection)
        with connection.cursor() as cursor:
            cursor.execute(f"SELECT id FROM {SETTLEMENT_TABLE} WHERE id = %s", (row_id,))
            if not cursor.fetchone():
                raise ValueError("记录不存在")
            cursor.execute(f"DELETE FROM {SETTLEMENT_TABLE} WHERE id = %s", (row_id,))
        connection.commit()

    return {
        "deleted_id": str(row_id),
        "rows": _settlements_payload(request=request),
    }


def _import_settlements_payload(rows_value: Any, request: FastAPIRequest | None = None) -> dict[str, Any]:
    if not isinstance(rows_value, list) or not rows_value:
        raise ValueError("没有可导入的数据")

    normalized_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for index, raw in enumerate(rows_value, start=1):
        try:
            row = _normalize_settlement_row(raw if isinstance(raw, dict) else {})
            key = _settlement_identity_key(row)
            if key in seen:
                continue
            seen.add(key)
            normalized_rows.append(row)
        except ValueError as exc:
            errors.append({"row": index, "error": str(exc)})

    if not normalized_rows:
        if errors:
            first_errors = "；".join(f"第{item['row']}行：{item['error']}" for item in errors[:5])
            raise ValueError("没有有效行可导入；" + first_errors)
        raise ValueError("没有有效行可导入")

    with _connect() as connection:
        _ensure_settlement_table(connection)
        with connection.cursor() as cursor:
            known_product_codes = _settlement_product_code_map(cursor)
        for row in normalized_rows:
            if not row["product_code"]:
                row["product_code"] = known_product_codes.get(_settlement_product_name_key(row.get("product_name"))) or None
        existing_keys: set[tuple[str, str, str, str]] = set()
        key_placeholders = ", ".join(["(%s, %s, %s, %s)"] * len(normalized_rows))
        key_params: list[Any] = []
        for row in normalized_rows:
            key_params.extend(_settlement_identity_key(row))
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT id, settlement_date, ip_name, product_code, article_type
                FROM {SETTLEMENT_TABLE}
                WHERE (settlement_date, ip_name, product_code, article_type) IN ({key_placeholders})
                """,
                tuple(key_params),
            )
            existing_ids: dict[tuple[str, str, str, str], int] = {}
            for row in cursor.fetchall():
                key = _settlement_identity_key(row)
                existing_keys.add(key)
                existing_ids.setdefault(key, int(row.get("id") or 0))

        placeholders = ", ".join(["%s"] * len(SETTLEMENT_DB_COLUMNS))
        update_columns = [column for column in SETTLEMENT_DB_COLUMNS if column not in {"settlement_date", "ip_name", "product_code", "article_type"}]
        update_sql = _settlement_import_update_sql(update_columns)
        with connection.cursor() as cursor:
            for row in normalized_rows:
                existing_id = existing_ids.get(_settlement_identity_key(row))
                if existing_id:
                    cursor.execute(
                        f"""
                        INSERT INTO {SETTLEMENT_TABLE} (id, {", ".join(SETTLEMENT_DB_COLUMNS)})
                        VALUES (%s, {placeholders})
                        ON DUPLICATE KEY UPDATE {update_sql}
                        """,
                        (existing_id, *tuple(row[column] for column in SETTLEMENT_DB_COLUMNS)),
                    )
                    continue
                cursor.execute(
                    f"""
                    INSERT INTO {SETTLEMENT_TABLE} ({", ".join(SETTLEMENT_DB_COLUMNS)})
                    VALUES ({placeholders})
                    ON DUPLICATE KEY UPDATE {update_sql}
                    """,
                    tuple(row[column] for column in SETTLEMENT_DB_COLUMNS),
                )
        connection.commit()

    inserted_count = sum(_settlement_identity_key(row) not in existing_keys for row in normalized_rows)
    updated_count = len(normalized_rows) - inserted_count
    return {
        "received_count": len(rows_value),
        "valid_count": len(normalized_rows),
        "inserted_count": inserted_count,
        "updated_count": updated_count,
        "error_count": len(errors),
        "errors": errors[:50],
        "rows": _settlements_payload(request=request),
    }


def _fill_settlement_fans_count_payload(ids_value: Any, request: FastAPIRequest | None = None) -> dict[str, Any]:
    if not isinstance(ids_value, list) or not ids_value:
        raise ValueError("请选择需要补全粉丝数的记录")

    row_ids: list[int] = []
    seen: set[int] = set()
    for value in ids_value:
        row_id = _parse_int(value)
        if row_id is None or row_id <= 0 or row_id in seen:
            continue
        seen.add(row_id)
        row_ids.append(row_id)
    if not row_ids:
        raise ValueError("没有有效的记录 ID")

    placeholders = ", ".join(["%s"] * len(row_ids))
    metrics_source_sql = f"""
        SELECT
          metric_date,
          TRIM(platform) COLLATE utf8mb4_unicode_ci AS platform_key,
          TRIM(kol_name) COLLATE utf8mb4_unicode_ci AS kol_name_key,
          MAX(fans_count) AS fans_count
        FROM {KOL_DAILY_METRICS_TABLE}
        WHERE fans_count IS NOT NULL
        GROUP BY metric_date, TRIM(platform) COLLATE utf8mb4_unicode_ci, TRIM(kol_name) COLLATE utf8mb4_unicode_ci
    """
    with _connect() as connection:
        _ensure_settlement_table(connection)
        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                SELECT
                  s.id,
                  s.settlement_date,
                  s.delivery_platform,
                  s.ip_name,
                  m.fans_count AS source_fans_count
                FROM {SETTLEMENT_TABLE} AS s
                LEFT JOIN ({metrics_source_sql}) AS m
                  ON m.metric_date = s.settlement_date
                 AND m.platform_key = TRIM(s.delivery_platform) COLLATE utf8mb4_unicode_ci
                 AND m.kol_name_key = TRIM(s.ip_name) COLLATE utf8mb4_unicode_ci
                WHERE s.id IN ({placeholders})
                """,
                tuple(row_ids),
            )
            target_rows = cursor.fetchall()

            matched_ids = [_parse_int(row.get("id")) for row in target_rows if row.get("source_fans_count") is not None]
            unmatched_ids = [_parse_int(row.get("id")) for row in target_rows if row.get("source_fans_count") is None]
            matched_ids = [row_id for row_id in matched_ids if row_id is not None]
            unmatched_ids = [row_id for row_id in unmatched_ids if row_id is not None]

            cursor.execute(
                f"""
                UPDATE {SETTLEMENT_TABLE} AS s
                JOIN ({metrics_source_sql}) AS m
                  ON m.metric_date = s.settlement_date
                 AND m.platform_key = TRIM(s.delivery_platform) COLLATE utf8mb4_unicode_ci
                 AND m.kol_name_key = TRIM(s.ip_name) COLLATE utf8mb4_unicode_ci
                SET s.fans_count = m.fans_count
                WHERE s.id IN ({placeholders})
                """,
                tuple(row_ids),
            )
            updated_count = cursor.rowcount
            connection.commit()

    return {
        "requested_count": len(row_ids),
        "target_count": len(target_rows),
        "matched_count": len(matched_ids),
        "updated_count": updated_count,
        "unmatched_count": len(unmatched_ids),
        "unmatched_ids": unmatched_ids[:50],
        "rows": _settlements_payload(request=request),
    }


def _settlement_export_number(value: Any) -> float:
    text = _clean_text(value).replace(",", "")
    if not text:
        return 0.0
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else 0.0


def _settlement_export_stat_value(field: str, rows: list[dict[str, Any]]) -> str:
    value = sum(_settlement_export_number(row.get(field)) for row in rows)
    if field in SETTLEMENT_EXPORT_MONEY_FIELDS:
        return f"{value:.2f}"
    return str(int(round(value)))


def _settlement_export_stats_row(fields: list[str], rows: list[dict[str, Any]]) -> list[Any]:
    values: list[Any] = []
    for index, field in enumerate(fields):
        if index == 0:
            values.append("SUM")
        elif field in SETTLEMENT_EXPORT_NUMERIC_FIELDS:
            values.append(_settlement_export_stat_value(field, rows))
        else:
            values.append("")
    return values


def _settlements_template_xlsx() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "社区大V业务看板模板"
    fields = SETTLEMENT_IMPORT_FIELDS

    worksheet.append([SETTLEMENT_EXPORT_LABELS.get(field, field) for field in fields])
    sample_row = {
        "date": date.today().strftime("%Y-%m-%d"),
        "partner": "示例合作方",
        "deliveryPlatform": "理财通",
        "product": "示例产品",
        "productCode": "000001",
        "ipName": "示例IP",
        "fansCount": "",
        "articleType": "加仓贴",
        "fee": "1000",
        "creatorFee": "500",
        "kolType": "外部",
        "buyAmount": "5000",
        "link": "https://example.com/post/unique-url",
        "articleTitle": "示例文章标题",
        "screenshot": "http://192.168.1.30:8898/captures/example/page_000.png",
        "readCount": "",
        "commentCount": "",
        "likeCount": "",
        "partnerPaymentStatus": "未打款",
        "creatorSettlementStatus": "未结算",
        "notes": "日期 + IP名称 + 代码 + 文章类型用于去重；产品仅用于展示，不再参与唯一键",
    }
    worksheet.append([sample_row.get(field, "") for field in fields])

    header_fill = PatternFill("solid", fgColor="EEF2F7")
    header_font = Font(bold=True, color="344054")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_cells in worksheet.iter_rows(min_row=2):
        for column_index, cell in enumerate(row_cells, start=1):
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fields[column_index - 1] == "productCode":
                cell.value = _clean_settlement_product_code(cell.value)
                cell.number_format = "@"

    for column_index, field in enumerate(fields, start=1):
        column_letter = get_column_letter(column_index)
        if field in {"link", "screenshot", "articleTitle", "notes", "product"}:
            worksheet.column_dimensions[column_letter].width = 38
        elif field == "productCode":
            worksheet.column_dimensions[column_letter].width = 14
        elif field in SETTLEMENT_EXPORT_NUMERIC_FIELDS:
            worksheet.column_dimensions[column_letter].width = 14
        else:
            worksheet.column_dimensions[column_letter].width = 18

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _read_export_image_bytes(reference: str) -> bytes | None:
    reference = _clean_text(reference)
    if not reference:
        return None
    parsed = urlparse(reference)
    try:
        if parsed.scheme in {"http", "https"}:
            url_request = UrlRequest(reference, headers={"User-Agent": "easy-viewer/1.0"})
            with urlopen(url_request, timeout=8) as response:
                return response.read(8 * 1024 * 1024)
        path = Path(reference)
        if path.exists() and path.is_file():
            return path.read_bytes()
    except Exception:
        return None
    return None


def _xlsx_image_from_reference(reference: str) -> tuple[Any, io.BytesIO] | None:
    image_bytes = _read_export_image_bytes(reference)
    if not image_bytes:
        return None
    try:
        from PIL import Image as PILImage
        from openpyxl.drawing.image import Image as XlsxImage

        image = PILImage.open(io.BytesIO(image_bytes))
        image.load()
        if image.mode not in {"RGB", "RGBA"}:
            image = image.convert("RGB")
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        buffer.seek(0)
        xlsx_image = XlsxImage(buffer)
        xlsx_image.width = image.width
        xlsx_image.height = image.height
        return xlsx_image, buffer
    except Exception:
        return None


def _settlements_xlsx_payload(payload: dict[str, Any]) -> bytes:
    rows_value = payload.get("rows")
    fields_value = payload.get("fields")
    if not isinstance(rows_value, list):
        rows_value = []
    if not isinstance(fields_value, list) or not fields_value:
        fields_value = SETTLEMENT_IMPORT_FIELDS

    allowed_fields = set(SETTLEMENT_IMPORT_FIELDS)
    fields = [str(field) for field in fields_value if str(field) in allowed_fields]
    if not fields:
        fields = SETTLEMENT_IMPORT_FIELDS
    rows = [row for row in rows_value if isinstance(row, dict)]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "社区大V业务看板"
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="EEF2F7")
    stat_fill = PatternFill("solid", fgColor="F8FAFC")
    header_font = Font(bold=True, color="344054")
    stat_font = Font(bold=True, color="102A43")

    worksheet.append([SETTLEMENT_EXPORT_LABELS.get(field, field) for field in fields])
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    image_buffers: list[io.BytesIO] = []
    screenshot_column = fields.index("screenshot") + 1 if "screenshot" in fields else 0
    max_screenshot_width = 0
    for row_index, row in enumerate(rows, start=2):
        values = [row.get(field, "") for field in fields]
        worksheet.append(values)
        for column_index, field in enumerate(fields, start=1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(vertical="center", wrap_text=field in {"articleTitle", "notes"})
            if field == "productCode":
                cell.value = _clean_settlement_product_code(cell.value)
                cell.number_format = "@"
            elif field in SETTLEMENT_EXPORT_NUMERIC_FIELDS and cell.value not in ("", None):
                cell.value = _settlement_export_number(cell.value)
                cell.number_format = "0.00" if field in SETTLEMENT_EXPORT_MONEY_FIELDS else "0"
        if screenshot_column:
            reference = _clean_text(row.get("screenshot"))
            image_result = _xlsx_image_from_reference(reference)
            if image_result:
                image, buffer = image_result
                image_buffers.append(buffer)
                cell_ref = f"{get_column_letter(screenshot_column)}{row_index}"
                worksheet.cell(row=row_index, column=screenshot_column).value = ""
                worksheet.add_image(image, cell_ref)
                max_screenshot_width = max(max_screenshot_width, image.width)
                worksheet.row_dimensions[row_index].height = max(90, min(409, image.height * 0.75 + 8))

    stat_row_index = len(rows) + 2
    worksheet.append(_settlement_export_stats_row(fields, rows))
    for cell in worksheet[stat_row_index]:
        cell.fill = stat_fill
        cell.font = stat_font
        cell.alignment = Alignment(horizontal="right" if isinstance(cell.value, (int, float)) else "left", vertical="center")

    for column_index, field in enumerate(fields, start=1):
        column_letter = get_column_letter(column_index)
        if field in {"link", "articleTitle", "product"}:
            worksheet.column_dimensions[column_letter].width = 42
        elif field == "productCode":
            worksheet.column_dimensions[column_letter].width = 14
        elif field == "screenshot":
            worksheet.column_dimensions[column_letter].width = min(255, max(58, max_screenshot_width / 7 if max_screenshot_width else 58))
        elif field in SETTLEMENT_EXPORT_NUMERIC_FIELDS:
            worksheet.column_dimensions[column_letter].width = 14
        else:
            worksheet.column_dimensions[column_letter].width = 18

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def _parse_dates(value: Any) -> list[date]:
    dates: list[date] = []
    for item in str(value or "").split(","):
        parsed = _parse_date(item.strip())
        if parsed and parsed not in dates:
            dates.append(parsed)
    return dates


def _parse_int(value: Any, default: int | None = None) -> int | None:
    if value in (None, ""):
        return default
    try:
        return int(str(value))
    except ValueError:
        return default


def _datetime_text(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value or "")


def _kol_columns() -> list[tuple[str, str]]:
    return [
        ("日期", "metric_date"),
        ("Title/大V名称", "kol_name"),
        ("平台", "platform"),
        ("类型", "kol_type"),
        ("第几群", "group_name"),
        ("主页", "homepage_url"),
        ("粉丝数", "fans_count"),
        ("增粉数", "growth_count"),
        ("阅读数", "read_count"),
        ("24h发文", "post_count_24h"),
        ("更新时间", "updated_at"),
        ("备注", "remark"),
        ("错误", "writeback_error"),
    ]


def _kol_normalize_filters(params: dict[str, Any]) -> dict[str, Any]:
    limit = _parse_int(params.get("limit"), 500) or 500
    sort = str(params.get("sort") or "base_id")
    if sort not in KOL_SORT_OPTIONS:
        sort = "base_id"
    missing = str(params.get("missing") or "")
    if missing not in KOL_MISSING_OPTIONS:
        missing = ""
    metric_dates = _parse_dates(params.get("date"))
    return {
        "metric_date": metric_dates[0] if len(metric_dates) == 1 else None,
        "metric_dates": metric_dates,
        "platform": str(params.get("platform") or "").strip(),
        "kol_type": str(params.get("kol_type") or "").strip(),
        "missing": missing,
        "q": str(params.get("q") or "").strip(),
        "limit": min(max(limit, 1), KOL_MAX_LIMIT),
        "sort": sort,
    }


def _kol_where_clause(filters: dict[str, Any]) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    args: list[Any] = []
    metric_dates = filters.get("metric_dates") or []
    if metric_dates:
        placeholders = ", ".join(["%s"] * len(metric_dates))
        clauses.append(f"m.metric_date IN ({placeholders})")
        args.extend(metric_dates)
    if filters["platform"]:
        clauses.append("m.platform = %s")
        args.append(filters["platform"])
    if filters["kol_type"]:
        clauses.append("COALESCE(NULLIF(b.kol_type, ''), '未匹配') = %s")
        args.append(filters["kol_type"])
    if filters["missing"] == "fans_empty":
        clauses.append("m.fans_count IS NULL")
    elif filters["missing"] == "growth_empty":
        clauses.append("m.growth_count IS NULL")
    elif filters["missing"] == "fans_or_growth_empty":
        clauses.append("(m.fans_count IS NULL OR m.growth_count IS NULL)")
    elif filters["missing"] == "fans_and_growth_empty":
        clauses.append("(m.fans_count IS NULL AND m.growth_count IS NULL)")
    if filters["q"]:
        clauses.append("(m.kol_name LIKE %s OR b.group_name LIKE %s)")
        like = f"%{filters['q']}%"
        args.extend([like, like])
    return ("WHERE " + " AND ".join(clauses), args) if clauses else ("", args)


def _kol_options(connection: Any) -> dict[str, Any]:
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT DISTINCT m.metric_date
            FROM {KOL_DAILY_METRICS_TABLE} AS m
            ORDER BY m.metric_date DESC
            """
        )
        dates = [_date_text(row.get("metric_date")) for row in cursor.fetchall()]

        cursor.execute(
            f"""
            SELECT DISTINCT m.platform
            FROM {KOL_DAILY_METRICS_TABLE} AS m
            ORDER BY m.platform
            """
        )
        platforms = [str(row.get("platform") or "") for row in cursor.fetchall()]

        cursor.execute(
            f"""
            SELECT DISTINCT COALESCE(NULLIF(b.kol_type, ''), '未匹配') AS kol_type
            FROM {KOL_DAILY_METRICS_TABLE} AS m
            LEFT JOIN {KOL_BASE_PROFILE_TABLE} AS b
              ON b.kol_name = m.kol_name
             AND b.platform = m.platform
            ORDER BY kol_type
            """
        )
        kol_types = [str(row.get("kol_type") or "") for row in cursor.fetchall()]

    return {"dates": dates, "platforms": platforms, "kol_types": kol_types}


def _kol_summary(connection: Any, filters: dict[str, Any]) -> dict[str, int]:
    where_sql, args = _kol_where_clause(filters)
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                COUNT(*) AS total_rows,
                COUNT(DISTINCT m.metric_date) AS date_count,
                COUNT(DISTINCT m.kol_name) AS kol_count,
                SUM(m.fans_count IS NOT NULL) AS fans_rows,
                SUM(m.growth_count IS NOT NULL) AS growth_rows,
                SUM(m.read_count IS NOT NULL) AS read_rows,
                SUM(m.post_count_24h IS NOT NULL) AS post_rows,
                SUM(COALESCE(NULLIF(b.kol_type, ''), '未匹配') = '内部') AS internal_rows,
                SUM(b.id IS NULL) AS unmatched_base_rows
            FROM {KOL_DAILY_METRICS_TABLE} AS m
            LEFT JOIN {KOL_BASE_PROFILE_TABLE} AS b
              ON b.kol_name = m.kol_name
             AND b.platform = m.platform
            {where_sql}
            """,
            args,
        )
        row = cursor.fetchone() or {}
    return {key: int(value or 0) for key, value in row.items()}


def _kol_row_remark(row: dict[str, Any]) -> str:
    payload = _json_value(row.get("source_payload_json"), {})
    warning = str(payload.get("quality_warning") or "").strip() if isinstance(payload, dict) else ""
    if warning:
        return warning
    if isinstance(payload, dict) and payload.get("nickname_mismatch"):
        return "昵称不一致"
    return ""


def _kol_api_row(row: dict[str, Any]) -> dict[str, Any]:
    output = {
        "metric_date": _date_text(row.get("metric_date")),
        "kol_name": str(row.get("kol_name") or ""),
        "platform": str(row.get("platform") or ""),
        "homepage_url": str(row.get("homepage_url") or ""),
        "group_name": str(row.get("group_name") or ""),
        "kol_type": str(row.get("kol_type") or ""),
        "fans_count": row.get("fans_count"),
        "growth_count": row.get("growth_count"),
        "read_count": row.get("read_count"),
        "post_count_24h": row.get("post_count_24h"),
        "updated_at": _datetime_text(row.get("updated_at")),
        "writeback_error": str(row.get("writeback_error") or ""),
    }
    output["remark"] = _kol_row_remark(row)
    return output


def _kol_rows(connection: Any, filters: dict[str, Any]) -> list[dict[str, Any]]:
    where_sql, args = _kol_where_clause(filters)
    order_sql = KOL_SORT_OPTIONS.get(filters["sort"], KOL_SORT_OPTIONS["base_id"])
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                m.metric_date,
                m.kol_name,
                m.platform,
                b.homepage_url,
                b.group_name,
                COALESCE(NULLIF(b.kol_type, ''), '未匹配') AS kol_type,
                m.fans_count,
                m.growth_count,
                m.read_count,
                m.post_count_24h,
                m.source_payload_json,
                m.writeback_error,
                m.updated_at
            FROM {KOL_DAILY_METRICS_TABLE} AS m
            LEFT JOIN {KOL_BASE_PROFILE_TABLE} AS b
              ON b.kol_name = m.kol_name
             AND b.platform = m.platform
            {where_sql}
            ORDER BY {order_sql}
            LIMIT %s
            """,
            (*args, int(filters["limit"])),
        )
        return [_kol_api_row(row) for row in cursor.fetchall()]


def _kol_metrics_payload(params: dict[str, Any]) -> dict[str, Any]:
    filters = _kol_normalize_filters(params)
    with _connect() as connection:
        return {
            "source": "mysql",
            "tables": [KOL_DAILY_METRICS_TABLE, KOL_BASE_PROFILE_TABLE],
            "filters": {
                **filters,
                "metric_date": _date_text(filters["metric_date"]),
                "metric_dates": [_date_text(item) for item in filters["metric_dates"]],
            },
            "missing_options": KOL_MISSING_OPTIONS,
            "sort_options": list(KOL_SORT_OPTIONS),
            "options": _kol_options(connection),
            "summary": _kol_summary(connection, filters),
            "columns": _kol_columns(),
            "rows": _kol_rows(connection, filters),
        }


def _kol_read_source_url(value: Any = None) -> str:
    _load_default_env()
    return _clean_text(value) or os.environ.get("KOL_READ_SOURCE_URL", "").strip() or KOL_READ_SOURCE_DEFAULT_URL


def _kol_read_field_from_header(header: Any) -> str:
    text = _compact_header(header)
    if not text:
        return ""
    if "日期" in text:
        return "metric_date"
    if text in {"账号名称", "大V名称", "Title/大V名称", "Title大V名称"} or ("大V名称" in text) or text == "Title":
        return "kol_name"
    if "文章阅读数" in text or text in {"阅读数", "阅读量"}:
        return "read_count"
    if "标题" in text:
        return "article_title"
    return ""


def _kol_read_date_text(value: Any) -> str:
    text = _clean_text(value).replace("/", "-").replace(".", "-")
    if not text:
        raise ValueError("缺少日期")
    short_match = re.fullmatch(r"(\d{2})(\d{2})", text)
    if short_match:
        year = datetime.now().year
        return f"{year:04d}-{int(short_match.group(1)):02d}-{int(short_match.group(2)):02d}"
    try:
        return _clean_date(text)
    except ValueError:
        match = re.fullmatch(r"(\d{1,2})-(\d{1,2})", text)
        if not match:
            raise
        year = datetime.now().year
        return f"{year:04d}-{int(match.group(1)):02d}-{int(match.group(2)):02d}"


def _kol_read_rows_from_matrix(matrix: list[list[Any]]) -> list[dict[str, Any]]:
    for header_index, row in enumerate(matrix[:30]):
        fields = [_kol_read_field_from_header(item) for item in row]
        if {"metric_date", "kol_name", "read_count"}.issubset(set(fields)):
            indexes = {field: fields.index(field) for field in ("metric_date", "kol_name", "read_count")}
            title_index = fields.index("article_title") if "article_title" in fields else None
            parsed_rows: list[dict[str, Any]] = []
            last_metric_date: Any = ""
            for data_row in matrix[header_index + 1 :]:
                if not any(_clean_text(item) for item in data_row):
                    continue
                parsed_row = {field: data_row[index] if index < len(data_row) else "" for field, index in indexes.items()}
                if _clean_text(parsed_row.get("metric_date")):
                    last_metric_date = parsed_row.get("metric_date")
                if not _clean_text(parsed_row.get("metric_date")) and title_index is not None and title_index < len(data_row):
                    title_date = _kol_read_date_from_title(data_row[title_index])
                    if title_date:
                        parsed_row["metric_date"] = title_date
                        last_metric_date = title_date
                if not _clean_text(parsed_row.get("metric_date")) and last_metric_date:
                    parsed_row["metric_date"] = last_metric_date
                parsed_rows.append(parsed_row)
            _fill_missing_kol_read_dates(parsed_rows)
            return parsed_rows
    return []


def _kol_read_default_date(value: Any) -> str:
    try:
        return _kol_read_date_text(value)
    except ValueError:
        return ""


def _kol_read_date_from_title(value: Any) -> str:
    text = _clean_text(value)
    patterns = [
        r"(?<!\d)(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})(?:日)?",
        r"(?<!\d)(\d{1,2})\s*月\s*(\d{1,2})\s*日",
        r"(?<!\d)(\d{2})(\d{2})(?!\d)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        groups = match.groups()
        try:
            if len(groups) == 3:
                parsed = date(int(groups[0]), int(groups[1]), int(groups[2]))
            else:
                parsed = date(datetime.now().year, int(groups[0]), int(groups[1]))
            return parsed.isoformat()
        except ValueError:
            continue
    return ""


def _fill_missing_kol_read_dates(rows: list[dict[str, Any]]) -> None:
    next_date = ""
    for row in reversed(rows):
        current = _clean_text(row.get("metric_date"))
        if current:
            next_date = current
        elif next_date:
            row["metric_date"] = next_date
    last_date = ""
    for row in rows:
        current = _clean_text(row.get("metric_date"))
        if current:
            last_date = current
        elif last_date:
            row["metric_date"] = last_date


def _parse_kol_read_source_text(text: str) -> list[dict[str, Any]]:
    clean = html_lib.unescape(str(text or "").lstrip("\ufeff"))
    candidates: list[list[list[Any]]] = []
    if "\t" in clean:
        candidates.append(list(csv.reader(io.StringIO(clean), delimiter="\t")))
    try:
        dialect = csv.Sniffer().sniff(clean[:4096])
        candidates.append(list(csv.reader(io.StringIO(clean), dialect)))
    except csv.Error:
        candidates.append(list(csv.reader(io.StringIO(clean))))
    for matrix in candidates:
        rows = _kol_read_rows_from_matrix(matrix)
        if rows:
            return rows
    return []


def _json_from_jsonp(text: str) -> Any:
    clean = html_lib.unescape(str(text or "").strip())
    match = re.match(r"^[\w$.\[\]]+\((.*)\)\s*;?$", clean, re.S)
    if match:
        clean = match.group(1)
    return json.loads(clean)


def _walk_values(value: Any) -> list[Any]:
    values: list[Any] = []
    if isinstance(value, dict):
        for item in value.values():
            values.append(item)
            values.extend(_walk_values(item))
    elif isinstance(value, list):
        for item in value:
            values.append(item)
            values.extend(_walk_values(item))
    return values


def _first_kol_read_table_text(payload: Any) -> str:
    for value in _walk_values(payload):
        if isinstance(value, str) and "T文章阅读数" in value and "账号名称" in value:
            return value
    return ""


def _fetch_url_text(url: str, timeout: int = 25, headers: dict[str, str] | None = None) -> str:
    request_headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) easy-viewer/1.0",
        "Accept": "text/html,application/json,text/plain,*/*",
    }
    request_headers.update(headers or {})
    request = UrlRequest(
        url,
        headers=request_headers,
    )
    with urlopen(request, timeout=timeout) as response:
        content = response.read()
        charset = response.headers.get_content_charset() or "utf-8"
    return content.decode(charset, errors="replace")


def _tencent_doc_headers() -> dict[str, str]:
    _load_default_env()
    config = _load_tencent_doc_config_from_app_config()
    access_token = config.get("TENCENT_DOC_ACCESS_TOKEN", "").strip()
    client_id = config.get("TENCENT_DOC_CLIENT_ID", "").strip()
    open_id = config.get("TENCENT_DOC_OPEN_ID", "").strip()
    client_secret = config.get("TENCENT_DOC_CLIENT_SECRET", "").strip()
    if not (access_token and client_id and open_id):
        return {}
    expiry = _tencent_doc_access_token_expiry(access_token)
    if expiry and expiry <= datetime.now(timezone.utc).astimezone():
        message = f"腾讯文档 Access-Token 已过期（{expiry.strftime('%Y-%m-%d %H:%M:%S %z')}）"
        if client_secret:
            message += "，请更新 crawler_app.app_config 中的 TENCENT_DOC_ACCESS_TOKEN；当前已配置 TENCENT_DOC_CLIENT_SECRET，后续可以接入自动换 token"
        else:
            message += "，请更新 crawler_app.app_config 中的 TENCENT_DOC_ACCESS_TOKEN；当前 TENCENT_DOC_CLIENT_SECRET 为空，无法自动换 token"
        raise ValueError(message)
    return {
        "Access-Token": access_token,
        "Client-Id": client_id,
        "Open-Id": open_id,
        "Content-Type": "application/json",
    }


def _load_tencent_doc_config_from_app_config() -> dict[str, str]:
    keys = (
        "TENCENT_DOC_ACCESS_TOKEN",
        "TENCENT_DOC_CLIENT_ID",
        "TENCENT_DOC_OPEN_ID",
        "TENCENT_DOC_CLIENT_SECRET",
        "TENCENT_DOC_TOKEN_URL",
    )
    placeholders = ", ".join(["%s"] * len(keys))
    try:
        with _connect() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    f"""
                    SELECT config_key, config_value
                    FROM crawler_app.app_config
                    WHERE status = 'active'
                      AND config_key IN ({placeholders})
                      AND config_value IS NOT NULL
                      AND config_value <> ''
                    """,
                    keys,
                )
                rows = cursor.fetchall()
    except Exception:
        return {}
    return {
        str(row.get("config_key") or ""): str(row.get("config_value") or "")
        for row in rows
        if str(row.get("config_key") or "") in keys
    }


def _tencent_doc_access_token_expiry(access_token: str) -> datetime | None:
    parts = access_token.split(".")
    if len(parts) != 3:
        return None
    try:
        payload_text = base64.urlsafe_b64decode(parts[1] + "=" * (-len(parts[1]) % 4)).decode("utf-8", errors="replace")
        payload = json.loads(payload_text)
    except Exception:
        return None
    exp_value = payload.get("exp") if isinstance(payload, dict) else None
    try:
        if exp_value is None:
            return None
        return datetime.fromtimestamp(float(exp_value), tz=timezone.utc).astimezone()
    except (TypeError, ValueError, OSError):
        return None


def _tencent_openapi_json(path_or_url: str) -> Any:
    headers = _tencent_doc_headers()
    if not headers:
        raise ValueError("缺少腾讯文档 OpenAPI 配置")
    url = path_or_url if path_or_url.startswith("http") else f"https://docs.qq.com{path_or_url}"
    request = UrlRequest(url, headers={**headers, "Accept": "application/json"})
    with urlopen(request, timeout=30) as response:
        content = response.read()
    payload = _json_from_jsonp(content.decode("utf-8", errors="replace"))
    if isinstance(payload, dict):
        code = payload.get("code", payload.get("ret", 0))
        if code not in (0, "0", None):
            raise ValueError(str(payload.get("message") or payload.get("msg") or "腾讯文档 OpenAPI 请求失败"))
    return payload


def _qq_docs_id_and_tab(source_url: str) -> tuple[str, str]:
    parsed = urlparse(source_url)
    parts = [part for part in parsed.path.split("/") if part]
    doc_id = parts[-1] if parts else ""
    tab = (parse_qs(parsed.query).get("tab") or [""])[0]
    return doc_id, tab


def _matrix_from_payload(value: Any) -> list[list[Any]]:
    if isinstance(value, dict) and isinstance(value.get("gridData"), dict):
        matrix: list[list[Any]] = []
        for row in value["gridData"].get("rows") or []:
            if not isinstance(row, dict):
                continue
            matrix.append([_tencent_cell_text(cell) for cell in row.get("values") or []])
        return matrix
    if isinstance(value, dict):
        for key, item in value.items():
            if key in {"values", "cellValues", "data"} and isinstance(item, list) and all(isinstance(row, list) for row in item):
                return item
        for item in value.values():
            matrix = _matrix_from_payload(item)
            if matrix:
                return matrix
    elif isinstance(value, list):
        if value and all(isinstance(row, list) for row in value):
            return value
        for item in value:
            matrix = _matrix_from_payload(item)
            if matrix:
                return matrix
    return []


def _tencent_cell_text(cell: Any) -> Any:
    if not isinstance(cell, dict):
        return ""
    value = cell.get("cellValue")
    if value is None:
        return ""
    if not isinstance(value, dict):
        return value
    for key in ("text", "number", "formula", "boolValue"):
        if value.get(key) is not None:
            return value.get(key)
    time_value = value.get("time")
    if isinstance(time_value, dict):
        try:
            year = int(time_value.get("year") or 0)
            month = int(time_value.get("month") or 0)
            day = int(time_value.get("day") or 0)
            if year and month and day:
                return date(year, month, day).isoformat()
        except (TypeError, ValueError):
            return ""
    rich_text = value.get("richText")
    if isinstance(rich_text, list):
        return "".join(str(item.get("text") or "") for item in rich_text if isinstance(item, dict))
    return ""


def _text_values_from_payload(value: Any) -> list[str]:
    items: list[str] = []
    if isinstance(value, dict):
        for item in value.values():
            items.extend(_text_values_from_payload(item))
    elif isinstance(value, list):
        for item in value:
            items.extend(_text_values_from_payload(item))
    elif isinstance(value, str):
        items.append(value)
    return items


def _tencent_sheet_values_rows(source_url: str) -> list[dict[str, Any]]:
    doc_id, tab = _qq_docs_id_and_tab(source_url)
    if not doc_id or not tab:
        raise ValueError("腾讯文档链接缺少 fileId 或 tab")
    if not _tencent_doc_headers():
        return []
    metadata = _tencent_openapi_json(f"/openapi/spreadsheet/v3/files/{doc_id}?concise=1")
    sheets = metadata.get("properties") or [] if isinstance(metadata, dict) else []
    sheet = next((item for item in sheets if str(item.get("sheetId") or "") == tab), None)
    row_total = min(max(int((sheet or {}).get("rowTotal") or 1000), 1), 5000)
    column_total = min(max(int((sheet or {}).get("columnTotal") or 26), 1), 26)
    end_column = chr(ord("A") + column_total - 1)
    range_name = os.environ.get("KOL_READ_SOURCE_RANGE", "").strip() or f"A1:{end_column}{row_total}"
    payload = _tencent_openapi_json(f"/openapi/spreadsheet/v3/files/{doc_id}/{tab}/{quote(range_name, safe=':')}")
    rows = _kol_read_rows_from_matrix(_matrix_from_payload(payload))
    if rows:
        return rows
    return []


def _fetch_qq_docs_sheet_text(source_url: str) -> str:
    doc_id, tab = _qq_docs_id_and_tab(source_url)
    if not doc_id:
        raise ValueError("腾讯文档链接缺少文档 ID")
    html = _fetch_url_text(source_url)
    token_match = re.search(r"[?&]t=(\d+)", html) or re.search(r'"t"\s*:\s*"?(\d+)"?', html)
    params = {
        "id": doc_id,
        "normal": "1",
        "outformat": "1",
        "startrow": "0",
        "endrow": "5000",
        "callback": "easyViewerCallback",
    }
    if tab:
        params["tab"] = tab
    if token_match:
        params["t"] = token_match.group(1)
    payload = _json_from_jsonp(_fetch_url_text(f"https://docs.qq.com/dop-api/opendoc?{urlencode(params)}"))
    table_text = _first_kol_read_table_text(payload)
    if table_text:
        return table_text
    raise ValueError("未能从腾讯文档响应中识别表格内容")


def _fetch_kol_read_source_rows(source_url: str) -> list[dict[str, Any]]:
    parsed = urlparse(source_url)
    if parsed.netloc.endswith("docs.qq.com") and "/sheet/" in parsed.path:
        rows = _tencent_sheet_values_rows(source_url)
        if rows:
            return rows
        try:
            rows = _parse_kol_read_source_text(_fetch_qq_docs_sheet_text(source_url))
            if rows:
                return rows
        except Exception:
            pass
    rows = _parse_kol_read_source_text(_fetch_url_text(source_url))
    if not rows:
        raise ValueError("来源中没有识别到 日期 / 账号名称 / T文章阅读数 三列")
    return rows


def _normalize_kol_read_source_rows(rows: Any, default_date: Any = None) -> tuple[list[dict[str, Any]], dict[str, int]]:
    if not isinstance(rows, list):
        raise ValueError("补充阅读数来源为空")
    normalized: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    stats = {"source_rows": len(rows), "valid_rows": 0, "empty_read_rows": 0, "invalid_rows": 0, "duplicate_rows": 0, "out_of_range_rows": 0}
    fallback_date = _kol_read_default_date(default_date)
    lookback_days = _parse_int(os.environ.get("KOL_READ_SYNC_LOOKBACK_DAYS"), 183) or 183
    min_metric_date = date.today() - timedelta(days=max(lookback_days, 1))
    for raw in rows:
        row = raw if isinstance(raw, dict) else {}
        metric_date = _first_non_empty_value(row.get("metric_date"), row.get("date"), row.get("日期"), fallback_date)
        kol_name = _first_non_empty_value(row.get("kol_name"), row.get("account_name"), row.get("账号名称"), row.get("Title/大V名称"), row.get("大V名称"))
        read_count = _first_non_empty_value(row.get("read_count"), row.get("T文章阅读数"), row.get("阅读数"))
        read_number = _clean_int(read_count)
        if read_number is None:
            stats["empty_read_rows"] += 1
            continue
        try:
            date_text = _kol_read_date_text(metric_date)
        except ValueError:
            stats["invalid_rows"] += 1
            continue
        metric_date_value = _parse_date(date_text)
        if metric_date_value and metric_date_value < min_metric_date:
            stats["out_of_range_rows"] += 1
            continue
        name_text = _clean_text(kol_name)
        if not name_text:
            stats["invalid_rows"] += 1
            continue
        key = (date_text, name_text, KOL_READ_SOURCE_PLATFORM, KOL_READ_SOURCE_TYPE)
        if key in normalized:
            stats["duplicate_rows"] += 1
            normalized[key]["read_count"] += read_number
        else:
            normalized[key] = {
                "metric_date": date_text,
                "kol_name": name_text,
                "platform": KOL_READ_SOURCE_PLATFORM,
                "kol_type": KOL_READ_SOURCE_TYPE,
                "read_count": read_number,
            }
    stats["valid_rows"] = len(normalized)
    return list(normalized.values()), stats


def _fill_kol_read_count_payload(payload: dict[str, Any] | None = None) -> dict[str, Any]:
    payload = payload or {}
    source_url = _kol_read_source_url(payload.get("source_url"))
    raw_rows = payload.get("rows")
    if raw_rows is None:
        raw_rows = _fetch_kol_read_source_rows(source_url)
    default_date = payload.get("date")
    if isinstance(default_date, str) and "," in default_date:
        default_date = ""
    rows, stats = _normalize_kol_read_source_rows(raw_rows, default_date)
    matched_count = 0
    updated_count = 0
    with _connect() as connection:
        with connection.cursor() as cursor:
            for row in rows:
                cursor.execute(
                    f"""
                    SELECT COUNT(*) AS matched_rows
                    FROM {KOL_DAILY_METRICS_TABLE} AS m
                    JOIN {KOL_BASE_PROFILE_TABLE} AS b
                      ON b.kol_name = m.kol_name
                     AND b.platform = m.platform
                    WHERE m.metric_date = %s
                      AND m.kol_name = %s
                      AND m.platform = %s
                      AND COALESCE(NULLIF(b.kol_type, ''), '未匹配') = %s
                    """,
                    (row["metric_date"], row["kol_name"], row["platform"], row["kol_type"]),
                )
                matched_count += int((cursor.fetchone() or {}).get("matched_rows") or 0)
                cursor.execute(
                    f"""
                    UPDATE {KOL_DAILY_METRICS_TABLE} AS m
                    JOIN {KOL_BASE_PROFILE_TABLE} AS b
                      ON b.kol_name = m.kol_name
                     AND b.platform = m.platform
                    SET m.read_count = %s,
                        m.updated_at = CURRENT_TIMESTAMP
                    WHERE m.metric_date = %s
                      AND m.kol_name = %s
                      AND m.platform = %s
                      AND COALESCE(NULLIF(b.kol_type, ''), '未匹配') = %s
                    """,
                    (row["read_count"], row["metric_date"], row["kol_name"], row["platform"], row["kol_type"]),
                )
                updated_count += int(getattr(cursor, "rowcount", 0) or 0)
        connection.commit()
    return {
        "source": source_url,
        "platform": KOL_READ_SOURCE_PLATFORM,
        "kol_type": KOL_READ_SOURCE_TYPE,
        "source_rows": stats["source_rows"],
        "valid_rows": stats["valid_rows"],
        "empty_read_rows": stats["empty_read_rows"],
        "invalid_rows": stats["invalid_rows"],
        "duplicate_rows": stats["duplicate_rows"],
        "out_of_range_rows": stats["out_of_range_rows"],
        "matched_count": matched_count,
        "updated_count": updated_count,
    }


def _kol_excel_value(value: Any) -> Any:
    return "" if value is None else value


def _hot_fund_columns() -> list[tuple[str, str]]:
    return [
        ("日期", "snapshot_date"),
        ("来源App", "source_app"),
        ("排名", "rank_no"),
        ("基金代码", "fund_code"),
        ("基金名称", "fund_name"),
        ("近一年收益率", "change_text"),
        ("截图", "screenshot_url"),
    ]


def _hot_fund_normalize_filters(params: dict[str, Any]) -> dict[str, Any]:
    limit = _parse_int(params.get("limit"), 200) or 200
    return {
        "snapshot_date": _parse_date(str(params.get("date") or "").strip()),
        "source_app": _clean_text(params.get("app") or params.get("source_app")),
        "limit": min(max(limit, 1), 1000),
    }


def _hot_fund_options(connection: Any) -> dict[str, Any]:
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT DISTINCT snapshot_date
            FROM {ALIPAY_HOT_FUND_RANKINGS_TABLE}
            ORDER BY snapshot_date DESC
            """
        )
        dates = [_date_text(row.get("snapshot_date")) for row in cursor.fetchall()]
        cursor.execute(
            f"""
            SELECT DISTINCT source_app
            FROM {ALIPAY_HOT_FUND_RANKINGS_TABLE}
            WHERE source_app IS NOT NULL
              AND source_app <> ''
            ORDER BY source_app
            """
        )
        apps = [str(row.get("source_app") or "") for row in cursor.fetchall()]
    return {"dates": dates, "apps": apps}


def _hot_fund_filter_sql(filters: dict[str, Any]) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    args: list[Any] = []
    if filters["snapshot_date"]:
        clauses.append("snapshot_date = %s")
        args.append(filters["snapshot_date"])
    if filters["source_app"]:
        clauses.append("source_app = %s")
        args.append(filters["source_app"])
    return ("WHERE " + " AND ".join(clauses), args) if clauses else ("", args)


def _hot_fund_rows(connection: Any, filters: dict[str, Any], request: FastAPIRequest | None = None) -> list[dict[str, Any]]:
    where_sql, args = _hot_fund_filter_sql(filters)
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT snapshot_date, source_app, rank_no, fund_code, fund_name, change_text, screenshot_path
            FROM {ALIPAY_HOT_FUND_RANKINGS_TABLE}
            {where_sql}
            ORDER BY snapshot_date DESC, rank_no ASC
            LIMIT %s
            """,
            (*args, int(filters["limit"])),
        )
        rows = []
        for row in cursor.fetchall():
            screenshot_path = str(row.get("screenshot_path") or "")
            rows.append(
                {
                    "snapshot_date": _date_text(row.get("snapshot_date")),
                    "source_app": str(row.get("source_app") or ""),
                    "rank_no": row.get("rank_no"),
                    "fund_code": str(row.get("fund_code") or ""),
                    "fund_name": str(row.get("fund_name") or ""),
                    "change_text": str(row.get("change_text") or ""),
                    "screenshot_path": screenshot_path,
                    "screenshot_url": _externalize_local_url(screenshot_path, request),
                }
            )
        return rows


def _hot_fund_summary(connection: Any, filters: dict[str, Any]) -> dict[str, int]:
    where_sql, args = _hot_fund_filter_sql(filters)
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                COUNT(*) AS total_rows,
                COUNT(DISTINCT snapshot_date) AS date_count,
                COUNT(DISTINCT source_app) AS app_count,
                SUM(screenshot_path IS NOT NULL AND screenshot_path <> '') AS screenshot_rows
            FROM {ALIPAY_HOT_FUND_RANKINGS_TABLE}
            {where_sql}
            """,
            args,
        )
        row = cursor.fetchone() or {}
    return {key: int(value or 0) for key, value in row.items()}


def _hot_fund_rankings_payload(params: dict[str, Any], request: FastAPIRequest | None = None) -> dict[str, Any]:
    filters = _hot_fund_normalize_filters(params)
    with _connect() as connection:
        return {
            "source": "mysql",
            "table": ALIPAY_HOT_FUND_RANKINGS_TABLE,
            "filters": {
                **filters,
                "snapshot_date": _date_text(filters["snapshot_date"]),
            },
            "options": _hot_fund_options(connection),
            "summary": _hot_fund_summary(connection, filters),
            "columns": _hot_fund_columns(),
            "rows": _hot_fund_rows(connection, filters, request),
        }


def _kol_metrics_xlsx(params: dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Missing dependency: pip install openpyxl") from exc

    payload = _kol_metrics_payload(params)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "kol_daily_metrics"
    columns = payload["columns"]
    worksheet.append([title for title, _key in columns])
    for row in payload["rows"]:
        worksheet.append([_kol_excel_value(row.get(key)) for _title, key in columns])

    header_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_cells in worksheet.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="center")

    widths = [12, 24, 10, 10, 10, 45, 12, 12, 12, 12, 20, 18, 36]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def create_app() -> FastAPI:
    _load_default_env()
    static_root = _viewer_root() / "post_viewer" / "static"
    capture_mounts = _capture_mounts()
    app = FastAPI(title="easy-viewer", version="0.1.0")

    @app.get("/health")
    def health() -> dict[str, str]:
        try:
            config = _mysql_config()
            return {"ok": "true", "source": "mysql", "database": str(config["database"]), "table": POST_TABLE}
        except Exception as exc:
            return {"ok": "false", "source": "mysql", "error": f"{type(exc).__name__}: {exc}"}

    @app.get("/", response_class=HTMLResponse)
    def index() -> str:
        return (static_root / "index.html").read_text(encoding="utf-8")

    @app.get("/rerun", response_class=HTMLResponse)
    def rerun_page() -> str:
        return (static_root / "index.html").read_text(encoding="utf-8")

    @app.get("/settlements", response_class=HTMLResponse)
    def settlements_page() -> str:
        return (static_root / "index.html").read_text(encoding="utf-8")

    @app.get("/kol-metrics", response_class=HTMLResponse)
    def kol_metrics_page() -> str:
        return (static_root / "index.html").read_text(encoding="utf-8")

    @app.get("/hot-funds", response_class=HTMLResponse)
    def hot_funds_page() -> str:
        return (static_root / "index.html").read_text(encoding="utf-8")

    @app.get("/post-production", response_class=HTMLResponse)
    def post_production_page() -> str:
        return (static_root / "index.html").read_text(encoding="utf-8")

    @app.get("/api/batches")
    def batches(limit: int = Query(default=30, ge=1, le=200)) -> list[dict[str, Any]]:
        try:
            return _batches_payload(limit)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.get("/api/dimensions")
    def dimensions() -> list[dict[str, Any]]:
        try:
            return _dimensions_payload()
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.get("/api/posts")
    def posts(trade_date: str = "", run_id: str = "", generated_at: str = "") -> dict[str, Any]:
        try:
            return _posts_payload(trade_date=trade_date.strip(), run_id=run_id.strip(), generated_at=generated_at.strip())
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.get("/api/post-production")
    def post_production(date: str = "", title: str = "", limit: int = Query(default=200, ge=1, le=1000)) -> dict[str, Any]:
        try:
            return _publish_tasks_payload({"date": date, "title": title, "limit": limit})
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.get("/api/public-base")
    def public_base(request: FastAPIRequest) -> dict[str, str]:
        scheme, host = _request_public_base(request) or (request.url.scheme, request.url.netloc)
        return {"base_url": f"{scheme}://{host}", "host": host}

    @app.post("/api/rerun-posts")
    def rerun_posts(payload: dict[str, Any]) -> dict[str, Any]:
        try:
            return _rerun_post_tasks_payload(payload.get("post_urls") or payload.get("urls") or "")
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.get("/api/settlement-autofill")
    def settlement_autofill(request: FastAPIRequest, post_url: str = Query(default="", min_length=1)) -> dict[str, Any]:
        try:
            return _settlement_autofill_payload(post_url, request)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.get("/api/settlements")
    def settlements(request: FastAPIRequest, limit: int = Query(default=500, ge=1, le=2000)) -> list[dict[str, Any]]:
        try:
            return _settlements_payload(limit, request)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.post("/api/settlements/fill-fans-count")
    def fill_settlement_fans_count(request: FastAPIRequest, payload: dict[str, Any]) -> dict[str, Any]:
        try:
            return _fill_settlement_fans_count_payload(payload.get("ids"), request)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.post("/api/settlements/update-engagement")
    def update_settlement_engagement(request: FastAPIRequest, payload: dict[str, Any]) -> dict[str, Any]:
        try:
            return _update_settlement_engagement_payload(payload, request)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.delete("/api/settlements/{settlement_id}")
    def delete_settlement(request: FastAPIRequest, settlement_id: int) -> dict[str, Any]:
        try:
            return _delete_settlement_payload(settlement_id, request)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.post("/api/settlements/export.xlsx")
    def export_settlements_xlsx(payload: dict[str, Any]) -> Response:
        try:
            data = _settlements_xlsx_payload(payload)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc
        filename = f"kol_business_settlements_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.get("/api/settlements/template.xlsx")
    def settlement_template_xlsx() -> Response:
        try:
            data = _settlements_template_xlsx()
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc
        return Response(
            data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="kol_business_settlements_template.xlsx"'},
        )

    @app.get("/api/kol-metrics")
    def kol_metrics(
        date: str = "",
        platform: str = "",
        kol_type: str = "",
        missing: str = "",
        q: str = "",
        sort: str = "base_id",
        limit: int = Query(default=500, ge=1, le=KOL_MAX_LIMIT),
    ) -> dict[str, Any]:
        try:
            return _kol_metrics_payload(
                {
                    "date": date,
                    "platform": platform,
                    "kol_type": kol_type,
                    "missing": missing,
                    "q": q,
                    "sort": sort,
                    "limit": limit,
                }
            )
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.get("/api/kol-metrics/export.xlsx")
    def kol_metrics_export(
        date: str = "",
        platform: str = "",
        kol_type: str = "",
        missing: str = "",
        q: str = "",
        sort: str = "base_id",
        limit: int = Query(default=500, ge=1, le=KOL_MAX_LIMIT),
    ) -> Response:
        try:
            data = _kol_metrics_xlsx(
                {
                    "date": date,
                    "platform": platform,
                    "kol_type": kol_type,
                    "missing": missing,
                    "q": q,
                    "sort": sort,
                    "limit": limit,
                }
            )
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc
        filename = f"kol_daily_metrics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.post("/api/kol-metrics/fill-read-count")
    def kol_metrics_fill_read_count(payload: dict[str, Any] | None = None) -> dict[str, Any]:
        try:
            return _fill_kol_read_count_payload(payload)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.get("/api/hot-funds")
    def hot_funds(
        request: FastAPIRequest,
        date: str = "",
        app: str = "",
        limit: int = Query(default=200, ge=1, le=1000),
    ) -> dict[str, Any]:
        try:
            return _hot_fund_rankings_payload({"date": date, "app": app, "limit": limit}, request)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.post("/api/settlements/import")
    def import_settlements(request: FastAPIRequest, payload: dict[str, Any]) -> dict[str, Any]:
        try:
            return _import_settlements_payload(payload.get("rows"), request)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    @app.post("/api/settlements/import-file")
    async def import_settlements_file(request: FastAPIRequest, file: UploadFile = File(...)) -> dict[str, Any]:
        try:
            content = await file.read()
            rows = _parse_settlement_import_file(content, file.filename or "")
            return _import_settlements_payload(rows, request)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"{type(exc).__name__}: {exc}") from exc

    app.mount("/static", StaticFiles(directory=static_root), name="static")
    for mount_path, capture_root in capture_mounts:
        mount_name = "captures_" + re.sub(r"[^a-zA-Z0-9_]+", "_", mount_path.strip("/"))
        app.mount(mount_path, StaticFiles(directory=capture_root), name=mount_name)
    return app


app = create_app()
